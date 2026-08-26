"""
Lacoste 26AW Takip Raporu v2
Kaynak: ANA TABLO sheet
Kurulum: pip install pandas openpyxl
Kullanim:
  python lacoste_rapor_v2.py --excel 26AW_TAKIP.xlsx
  python lacoste_rapor_v2.py --excel 26AW_TAKIP.xlsx --indir
"""
import json, sys, webbrowser, argparse
from datetime import datetime
from pathlib import Path
import pandas as pd

SCRIPT_DIR = Path(__file__).parent

KANALLAR = {
    "SSTEP": {
        "sip_haric":   ["Tum SSTEP"],
        "sip_dahil":   ["Tum SSTEP", "Bayi Devir Adet"],
        "mag_ytd":     "SSTEP TY SLSU(OMSHaric)",
        "mag_sh":      "SSTEP TY SLSU(OMSHaric) Son Hafta",
        "mag_mtd":     "Aylik TY SLSU(OMSHaric) SSTEP",
        "onl_ytd":     "SSTEP Online Satis",
        "onl_sh":      "SSTEP Online Son Hafta Satis",
        "onl_mtd":     "Online Aylik Satis SSTEP",
        "stok":        "SSTEP TY STKU (SonGun)",
        "ros":         "ROS",
        "sevk":        "Dagitim Sevk Tarihi SSTEP",
        "renk":        "#0057B8",
        "renk_ac":     "#e6f0fb",
        "emoji":       "SSTEP",
    },
    "Fashfed": {
        "sip_haric":   ["FASHFED P10"],
        "sip_dahil":   ["FASHFED P10"],
        "mag_ytd":     "TY SLSU(OMSHaric) Fashfed",
        "mag_sh":      "TY SLSU(OMSHaric) Fashfed Son Hafta",
        "mag_mtd":     "Aylik TY SLSU(OMSHaric) Fashfed",
        "onl_ytd":     "Online Satis Fashfed",
        "onl_sh":      "Online Son Hafta Satis Fashfed",
        "onl_mtd":     "Online Aylik Satis Fashfed",
        "stok":        "TY STKU (SonGun) Fashfed",
        "ros":         "ross",
        "sevk":        "Aktif Sevk Tarihi",
        "renk":        "#E8003D",
        "renk_ac":     "#fde8ee",
        "emoji":       "Fashfed",
    },
    "Intersport": {
        "sip_haric":   ["INTERSPORT P10"],
        "sip_dahil":   ["INTERSPORT P10"],
        "mag_ytd":     "TY SLSU(OMSHaric) Intersport",
        "mag_sh":      "Son Hafta TY SLSU(OMSHaric) Intersport",
        "mag_mtd":     "Aylik TY SLSU(OMSHaric) Intersport",
        "onl_ytd":     "Online Satis Intersport",
        "onl_sh":      "Online Son Hafta Satis Intersport",
        "onl_mtd":     "Online Aylik Satis Intersport",
        "stok":        "TY STKU (SonGun) Intersport",
        "ros":         None,
        "sevk":        "Aktif Sevk Tarihi",
        "renk":        "#FF6B00",
        "renk_ac":     "#fff0e6",
        "emoji":       "Intersport",
    },
    "Lacoste": {
        "sip_haric":   ["TR Sipariş Toplam"],
        "sip_dahil":   ["TR Sipariş Toplam", "TR Devir Adet"],
        "mag_ytd":     "TY SLSU(OMSHaric) Lacoste",
        "mag_sh":      "TY SLSU(OMSHaric) Lacoste Son Hafta",
        "mag_mtd":     "Aylik TY SLSU(OMSHaric) Lacoste",
        "onl_ytd":     "Online Satis Lacoste",
        "onl_sh":      None,
        "onl_mtd":     "Online Aylik Satis Lacoste",
        "stok":        "TY STKU (SonGun) Lacoste",
        "ros":         None,
        "sevk":        "Aktif Sevk Tarihi",
        "renk":        "#C8B400",
        "renk_ac":     "#fdf9e0",
        "emoji":       "Lacoste",
    },
}


def log(msg, sym="  "):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {sym} {msg}")


def safe(s):
    if isinstance(s, pd.Series):
        return pd.to_numeric(s, errors='coerce').fillna(0)
    return pd.Series(0)


def find_col(df, name):
    """Turkce karakterleri normalize ederek kolonu bul"""
    if name in df.columns:
        return name
    tr = [("c","c"),("C","C"),("g","g"),("G","G"),("i","i"),
          ("I","I"),("o","o"),("O","O"),("s","s"),("S","S"),("u","u"),("U","U")]
    norm_name = name
    for nt, en in [("c","c"),("g","g"),("i","i"),("o","o"),("s","s"),("u","u"),
                   ("C","C"),("G","G"),("I","I"),("O","O"),("S","S"),("U","U")]:
        pass
    # Basit normalize
    def normalize(t):
        for o,n in [("\xe7","c"),("\xc7","C"),("\u011f","g"),("\u011e","G"),
                    ("\u0131","i"),("\u0130","I"),("\xf6","o"),("\xd6","O"),
                    ("\u015f","s"),("\u015e","S"),("\xfc","u"),("\xdc","U")]:
            t = t.replace(o, n)
        return t.strip()
    norm_target = normalize(name)
    for c in df.columns:
        if normalize(c) == norm_target:
            return c
    return name


def oku_excel(excel_yolu: Path):
    log(f"Excel okunuyor: {excel_yolu.name}", ">>")
    xl = pd.ExcelFile(excel_yolu)
    sheet = next((s for s in xl.sheet_names if "ANA TABLO" in s.upper()), xl.sheet_names[0])
    log(f"Sheet: {sheet}", "i")

    # Satır 6'dan (index 6) kanal isimlerini oku
    raw = pd.read_excel(excel_yolu, sheet_name=sheet, header=None)
    kanal_row = raw.iloc[6] if len(raw) > 6 else pd.Series(dtype=object)
    IGNORE = {'', 'nan', '50 adet', 'adet'}
    kanallar_ham = []
    for v in kanal_row:
        vs = str(v).strip() if pd.notna(v) else ''
        if vs and vs.lower() not in IGNORE:
            try: float(vs); continue
            except: pass
            kanallar_ham.append(vs)
    log(f"Kanallar: {kanallar_ham}", "i")

    df = pd.read_excel(excel_yolu, sheet_name=sheet, header=8)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=["M.R"]).reset_index(drop=True)
    log(f"{len(df)} satir okundu", "OK")

    def fc(cols, *pats):
        cu = [c.upper() for c in cols]
        for p in pats:
            pu = p.upper()
            for i, c in enumerate(cu):
                if pu in c:
                    return cols[i]
        return None

    cols = list(df.columns)
    RENKLER = ['#0057B8', '#C8B400', '#E8003D', '#FF6B00', '#9B59B6', '#1ABC9C', '#E74C3C', '#2ECC71']

    # Dinamik KANALLAR oluştur
    global KANALLAR
    KANALLAR_YEN = {}
    for idx_k, kanal in enumerate(kanallar_ham):
        k_up = kanal.upper()
        sip_h = fc(cols, f'Tüm {kanal}', f'{kanal} P10', f'{k_up} SİPARİŞ', f'{k_up} SIPARIS')
        if not sip_h and k_up == 'LACOSTE':
            sip_h = fc(cols, 'TR Sipariş Toplam', 'TR Siparis Toplam', 'TR Siparış Toplam')
        if not sip_h:
            sip_h = fc(cols, f'{k_up} TOPLAM', f'TÜM {k_up}')
        if not sip_h:
            continue  # Sipariş kolonu bulunamadı, bu kanalı atla

        sip_d_extra = fc(cols, f'{kanal} Devir', 'Bayi Devir Adet', 'TR Devir Adet')
        mag_ytd = fc(cols, f'TY SLSU(OMSHaric) {kanal}', f'{kanal} TY SLSU(OMSHaric)', f'SSTEP TY SLSU(OMSHaric)' if k_up=='SSTEP' else '__')
        mag_sh  = fc(cols, f'TY SLSU(OMSHaric) {kanal} Son Hafta', f'{kanal} TY SLSU(OMSHaric) Son Hafta')
        mag_mtd = fc(cols, f'Aylık TY SLSU(OMSHaric) {kanal}', f'{kanal} Aylık TY SLSU', f'Aylik TY SLSU(OMSHaric) {kanal}')
        onl_ytd = fc(cols, f'Online Satış {kanal}', f'{kanal} Online Satış', f'{kanal} Online Satis')
        onl_sh  = fc(cols, f'Online Son Hafta Satış {kanal}', f'{kanal} Online Son Hafta')
        onl_mtd = fc(cols, f'Online Aylık Satış {kanal}', f'{kanal} Online Aylık', f'Online Aylik Satis {kanal}')
        stok    = fc(cols, f'TY STKU (SonGun) {kanal}', f'{kanal} TY STKU')
        ros_col = fc(cols, 'ROS', 'ross')
        psf_col = fc(cols, 'PSF')

        sevk = fc(cols, f'Dağıtım Sevk Tarihi {kanal}', f'{kanal} Sevk', 'Aktif Sevk Tarihi', 'Dagitim Sevk Tarihi')
        KANALLAR_YEN[kanal] = {
            "renk":    RENKLER[idx_k % len(RENKLER)],
            "sip_haric": [sip_h] if sip_h else [],
            "sip_dahil": [sip_h, sip_d_extra] if sip_d_extra and sip_h else ([sip_h] if sip_h else []),
            "mag_ytd":  mag_ytd or "",
            "mag_sh":   mag_sh  or "",
            "mag_mtd":  mag_mtd or "",
            "onl_ytd":  onl_ytd or "",
            "onl_sh":   onl_sh  or "",
            "onl_mtd":  onl_mtd or "",
            "stok":     stok    or "",
            "sevk":     sevk    or "",
        }
        log(f"  {kanal}: sip={sip_h}, mag={mag_ytd}, onl={onl_ytd}", "i")

    if KANALLAR_YEN:
        KANALLAR = KANALLAR_YEN
    else:
        log("Dinamik kanal tespiti başarısız, varsayılan kullanılıyor", "!")

    extras = {}
    for kanal, cfg in KANALLAR.items():
        sip_h_cols = [c for c in cfg["sip_haric"] if c and c in df.columns]
        sip_d_cols = [c for c in cfg["sip_dahil"] if c and c in df.columns]
        sh  = df[[c for c in sip_h_cols]].apply(pd.to_numeric,errors='coerce').fillna(0).sum(axis=1).clip(lower=0) if sip_h_cols else pd.Series(0,index=df.index)
        sd  = df[[c for c in sip_d_cols]].apply(pd.to_numeric,errors='coerce').fillna(0).sum(axis=1).clip(lower=0) if sip_d_cols else sh
        def gcol(key):
            col=cfg.get(key,"")
            return pd.to_numeric(df[col],errors='coerce').fillna(0).clip(lower=0) if col and col in df.columns else pd.Series(0,index=df.index)
        my=gcol("mag_ytd"); ms=gcol("mag_sh"); mm=gcol("mag_mtd")
        oy=gcol("onl_ytd"); os_=gcol("onl_sh"); om=gcol("onl_mtd")
        st=gcol("stok")
        strh=(my+oy)/sh.replace(0,float('nan'))*100
        strd=(my+oy)/sd.replace(0,float('nan'))*100
        # urun_listesi beklediği format
        extras[f"_sh_{kanal}"]   = sh
        extras[f"_sd_{kanal}"]   = sd
        extras[f"_my_{kanal}"]   = my
        extras[f"_oy_{kanal}"]   = oy
        extras[f"_ty_{kanal}"]   = my+oy
        extras[f"_ms_{kanal}"]   = ms
        extras[f"_os_{kanal}"]   = os_
        extras[f"_ts_{kanal}"]   = ms+os_
        extras[f"_st_{kanal}"]   = st
        extras[f"_ro_{kanal}"]   = pd.Series(0,index=df.index)
        extras[f"_strh_{kanal}"] = strh.fillna(0).round(1)
        extras[f"_strd_{kanal}"] = strd.fillna(0).round(1)
        extras[f"_mmtd_{kanal}"] = mm
        extras[f"_omtd_{kanal}"] = om
        extras[f"_tmtd_{kanal}"] = mm+om

    psf_c = fc(cols, "PSF")
    extras["_psf"] = pd.to_numeric(df[psf_c], errors='coerce').fillna(0) if psf_c and psf_c in df.columns else pd.Series(0, index=df.index)
    wms_c = fc(cols, "WMS")
    extras["_wms"] = pd.to_numeric(df[wms_c], errors='coerce').fillna(0).clip(lower=0) if wms_c and wms_c in df.columns else pd.Series(0, index=df.index)
    depo_c = fc(cols, "DEPO STOK")
    extras["_depo"] = pd.to_numeric(df[depo_c], errors='coerce').fillna(0).clip(lower=0) if depo_c and depo_c in df.columns else pd.Series(0, index=df.index)

    img_c = fc(cols, "Fotograf", "Image", "Resim", "IMG")
    extras["_img"] = df[img_c].fillna("") if img_c and img_c in df.columns else pd.Series("", index=df.index)

    return pd.concat([df, pd.DataFrame(extras, index=df.index)], axis=1)


def get_col_val(df, row, name, default=""):
    c = find_col(df, name)
    if c in df.columns:
        v = row.get(c, default)
        return str(v).strip() if pd.notna(v) else str(default)
    return str(default)



def _hoss_urun_js(J, rows):
    items = []
    for p in rows:
        kv_parts = []
        for kanal in KANALLAR:
            kv_parts.append(f"{kanal}_sip_h:{p.get(kanal+'_sip_h',0)}")
            kv_parts.append(f"{kanal}_top_ytd:{p.get(kanal+'_top_ytd',0)}")
            kv_parts.append(f"{kanal}_stok:{p.get(kanal+'_stok',0)}")
            kv_parts.append(f"{kanal}_str_h:{p.get(kanal+'_str_h',0)}")
        kv = ",".join(kv_parts)
        line = (
            "{" + f"code:{J(p['code'])},gender:{J(p['gender'])},"
            f"ana_kat:{J(p['ana_kat'])},ust_kat:{J(p['ust_kat'])},"
            f"alt_kat:{J(p['alt_kat'])},story:{J(p['story'])},"
            f"img:{J(p['img'])},{kv}" + "}"
        )
        items.append(line)
    return "[" + ",".join(items) + "]"


def hoss_analiz(df):
    """HOSS X Lacoste Kapsul Koleksiyonu analizi - tum kanallar icin"""
    mask = df["STORY"].astype(str).str.contains("HOSS", case=False, na=False)
    hoss = df[mask].copy()
    
    if hoss.empty:
        return {"urunler": [], "ozet_kat": [], "ozet_ust": [], "ozet_cin": []}
    
    def safe_s(s):
        return pd.to_numeric(s, errors="coerce").fillna(0)
    
    urunler = []
    for _, r in hoss.iterrows():
        url = str(r.get("URL","")).strip()
        if url in ("nan","None",""): url = ""
        row = {
            "code":     str(r["M.R"]).strip(),
            "gender":   str(r.get("CINSIYET","")).strip(),
            "ana_kat":  get_col_val(df, r, "ANA KATEGORI"),
            "ust_kat":  get_col_val(df, r, "UST KATEGORI"),
            "alt_kat":  get_col_val(df, r, "ALT KATEGORI"),
            "story":    str(r.get("STORY","")).strip(),
            "img":      url,
        }
        for kanal, cfg in KANALLAR.items():
            sh_c = f"_sh_{kanal}"; sd_c = f"_sd_{kanal}"
            my_c = f"_my_{kanal}"; oy_c = f"_oy_{kanal}"
            ty_c = f"_ty_{kanal}"; st_c = f"_st_{kanal}"
            sh = int(r.get(sh_c, 0))
            sd = int(r.get(sd_c, 0))
            my = int(r.get(my_c, 0))
            oy = int(r.get(oy_c, 0))
            ty = my + oy
            st = int(r.get(st_c, 0))
            row[f"{kanal}_sip_h"]   = sh
            row[f"{kanal}_sip_d"]   = sd
            row[f"{kanal}_mag_ytd"] = my
            row[f"{kanal}_onl_ytd"] = oy
            row[f"{kanal}_top_ytd"] = ty
            row[f"{kanal}_stok"]    = st
            row[f"{kanal}_str_h"]   = round(ty/sh*100,1) if sh>0 else 0
            row[f"{kanal}_str_d"]   = round(ty/sd*100,1) if sd>0 else 0
        urunler.append(row)
    
    def grp_str(sub, group_col, label_key):
        rows = []
        col = find_col(df, group_col)
        if col not in hoss.columns:
            return rows
        for val in sorted(hoss[col].dropna().unique()):
            gsub = hoss[hoss[col]==val]
            r2 = {label_key: str(val)}
            for kanal in KANALLAR:
                sh = int(gsub[f"_sh_{kanal}"].sum())
                sd = int(gsub[f"_sd_{kanal}"].sum())
                my = int(gsub[f"_my_{kanal}"].sum())
                oy = int(gsub[f"_oy_{kanal}"].sum())
                ty = my + oy
                st = int(gsub[f"_st_{kanal}"].sum())
                r2[f"{kanal}_sip_h"]   = sh
                r2[f"{kanal}_top_ytd"] = ty
                r2[f"{kanal}_stok"]    = st
                r2[f"{kanal}_str_h"]   = round(ty/sh*100,1) if sh>0 else 0
            rows.append(r2)
        return rows
    
    return {
        "urunler":   urunler,
        "ozet_kat":  grp_str(hoss, "ANA KATEGORI",  "kat"),
        "ozet_ust":  grp_str(hoss, "UST KATEGORI",  "ust"),
        "ozet_cin":  grp_str(hoss, "CINSIYET",      "cin"),
        "count":     len(hoss),
    }


def urun_listesi(df, kanal):
    mask = df[f"_sh_{kanal}"] > 0
    rows = []
    for _, r in df[mask].iterrows():
        url = str(r.get(find_col(df, "URL"), "")).strip()
        if url in ("nan","None",""): url = ""
        rows.append({
            "code":     str(r["M.R"]).strip(),
            "gender":   str(r.get("CINSIYET","")).strip(),
            "ana_kat":  get_col_val(df, r, "ANA KATEGORI"),
            "ust_kat":  get_col_val(df, r, "UST KATEGORI"),
            "alt_kat":  get_col_val(df, r, "ALT KATEGORI"),
            "aile":     (lambda c: str(r[c]).strip() if c in df.columns and pd.notna(r.get(c)) else "")(
                            next((c for c in df.columns if c.strip() in ("AYAKKABI AİLE","AYAKKABI AILE")), "AYAKKABI AILE")),
            "alt_aile": (lambda c: str(r[c]).strip() if c in df.columns and pd.notna(r.get(c)) else "")(
                            next((c for c in df.columns if c.strip() in ("AYAKKABI ALT AİLE","AYAKKABI ALT AILE")), "AYAKKABI ALT AILE")),
            "sip_grp":  get_col_val(df, r, "SIPARIS GRUBU"),
            "sip_h":    int(r[f"_sh_{kanal}"]),
            "sip_d":    int(r[f"_sd_{kanal}"]),
            "mag_ytd":  int(r[f"_my_{kanal}"]),
            "onl_ytd":  int(r[f"_oy_{kanal}"]),
            "top_ytd":  int(r[f"_ty_{kanal}"]),
            "mag_sh":   int(r[f"_ms_{kanal}"]),
            "onl_sh":   int(r[f"_os_{kanal}"]),
            "top_sh":   int(r[f"_ts_{kanal}"]),
            "stok":     int(r[f"_st_{kanal}"]),
            "ros":      float(round(r[f"_ro_{kanal}"],2)),
            "str_h":    float(r[f"_strh_{kanal}"]),
            "str_d":    float(r[f"_strd_{kanal}"]),
            "psf":      int(r.get("_psf", 0) or 0),
            "wms":      int(r.get("_wms", 0) or 0),
            "depo":     int(r.get("_depo", 0) or 0),
            "mag_mtd":  int(r.get(f"_mmtd_{kanal}", 0) or 0),
            "onl_mtd":  int(r.get(f"_omtd_{kanal}", 0) or 0),
            "top_mtd":  int(r.get(f"_tmtd_{kanal}", 0) or 0),
            "img":      url,
            "story":    get_col_val(df, r, "STORY"),
            "is_hoss":  "HOSS" in str(r.get("STORY","")).upper(),
        })
    return rows


def str_analiz(df, kanal):
    sh = int(df[f"_sh_{kanal}"].sum())
    sd = int(df[f"_sd_{kanal}"].sum())
    my = int(df[f"_my_{kanal}"].sum())
    oy = int(df[f"_oy_{kanal}"].sum())
    ty = my + oy
    st = int(df[f"_st_{kanal}"].sum())
    toplam = {
        "sip_h": sh, "sip_d": sd, "mag_ytd": my, "onl_ytd": oy,
        "top_ytd": ty, "stok": st,
        "str_h": round(ty/sh*100,1) if sh>0 else 0,
        "str_d": round(ty/sd*100,1) if sd>0 else 0,
    }

    kat_col = find_col(df, "ANA KATEGORI")
    kat_rows = []
    if kat_col in df.columns:
        for kat in sorted(df[kat_col].dropna().unique()):
            sub = df[df[kat_col]==kat]
            ksh = int(sub[f"_sh_{kanal}"].sum())
            ksd = int(sub[f"_sd_{kanal}"].sum())
            kmy = int(sub[f"_my_{kanal}"].sum())
            koy = int(sub[f"_oy_{kanal}"].sum())
            kty = kmy + koy
            kat_rows.append({
                "kat": str(kat), "sip_h": ksh, "sip_d": ksd,
                "mag_ytd": kmy, "onl_ytd": koy, "top_ytd": kty,
                "str_h": round(kty/ksh*100,1) if ksh>0 else 0,
                "str_d": round(kty/ksd*100,1) if ksd>0 else 0,
            })
    # Cinsiyet bazlı STR
    cin_col = "CINSIYET"
    cin_rows = []
    if cin_col in df.columns:
        for cin in sorted(df[cin_col].dropna().unique()):
            sub = df[df[cin_col] == cin]
            csh = int(sub[f"_sh_{kanal}"].sum())
            csd = int(sub[f"_sd_{kanal}"].sum())
            cmy = int(sub[f"_my_{kanal}"].sum())
            coy = int(sub[f"_oy_{kanal}"].sum())
            cty = cmy + coy
            cin_rows.append({
                "cin": str(cin),
                "sip_h": csh, "sip_d": csd,
                "mag_ytd": cmy, "onl_ytd": coy, "top_ytd": cty,
                "str_h": round(cty/csh*100,1) if csh>0 else 0,
                "str_d": round(cty/csd*100,1) if csd>0 else 0,
            })

    # Aile bazlı STR
    # AYAKKABI AİLE kolonunu bul (Turkce I harfi dikkate alinarak)
    aile_col_name = next((c for c in df.columns
        if ("AYAKKABI" in c.upper() or "AYAKKABL" in c.upper())
        and "ALT" not in c.upper()
        and ("ALE" in c.upper() or "İLE" in c or "ILE" in c.upper())), None)
    if not aile_col_name:
        aile_col_name = next((c for c in df.columns if c.strip() in ("AYAKKABI AİLE", "AYAKKABI AILE")), None)
    aile_rows = []
    if aile_col_name and aile_col_name in df.columns:
        df_aile = df[(df[f"_sh_{kanal}"] > 0) & (~df[aile_col_name].astype(str).isin(["0","-","nan",""]))]
        for aile in sorted(df_aile[aile_col_name].dropna().unique()):
            sub = df_aile[df_aile[aile_col_name] == aile]
            ash = int(sub[f"_sh_{kanal}"].sum())
            asd = int(sub[f"_sd_{kanal}"].sum())
            amy = int(sub[f"_my_{kanal}"].sum())
            aoy = int(sub[f"_oy_{kanal}"].sum())
            aty = amy + aoy
            ast_ = int(sub[f"_st_{kanal}"].sum())
            aile_rows.append({
                "aile": str(aile), "sip_h": ash, "sip_d": asd,
                "mag_ytd": amy, "onl_ytd": aoy, "top_ytd": aty, "stok": ast_,
                "str_h": round(aty/ash*100,1) if ash>0 else 0,
                "str_d": round(aty/asd*100,1) if asd>0 else 0,
            })
        aile_rows.sort(key=lambda x: x["str_h"], reverse=True)

    return {"toplam": toplam, "kat": kat_rows, "cin": cin_rows, "aile": aile_rows}


def gitmeyen_df(df, kanal):
    cfg  = KANALLAR[kanal]
    sip  = df[f"_sh_{kanal}"]
    top  = df[f"_ty_{kanal}"]
    stok = df[f"_st_{kanal}"]
    depo = safe(df[find_col(df,"DEPO STOK")]) if find_col(df,"DEPO STOK") in df.columns else pd.Series(0, index=df.index)
    wms  = safe(df[find_col(df,"WMS")]) if find_col(df,"WMS") in df.columns else pd.Series(0, index=df.index)

    # Siparisi olan ve hic satilmamis ya da henuz mağazaya gitmemis
    mask = (sip > 0) & ((top == 0) | ((stok == 0) & (top == 0)))
    sub  = df[mask].copy()
    sub["_SIP"]  = sip[mask].values
    sub["_DEPO"] = depo[mask].values
    sub["_WMS"]  = wms[mask].values
    sub["_STOK"] = stok[mask].values

    def depo_durum(row):
        if row["_STOK"] > 0:
            return "MAGAZAYA GITMIS SATIS YOK"
        if row["_DEPO"] > 0:
            return "DEPODA VAR SEVK BEKLIYOR"
        if row["_WMS"] > 0:
            return "WMSTE VAR"
        return "DEPOYA GELMEMIS"

    def gidis_pct(row):
        return round(row["_STOK"] / row["_SIP"] * 100, 1) if row["_SIP"] > 0 else 0.0

    def gidis_durum(pct):
        if pct >= 100: return "TAMAMLANDI"
        if pct >= 75:  return "IYI"
        if pct >= 50:  return "AZ"
        if pct >= 25:  return "KRITIK"
        return "GITMEMIS"

    sub["Depo Durumu"] = sub.apply(depo_durum, axis=1)
    sub["Gidis Yuzde"] = sub.apply(gidis_pct, axis=1)
    sub["Gidis Durumu"] = sub["Gidis Yuzde"].apply(gidis_durum)

    sevk_c   = find_col(df, cfg.get("sevk","")) if cfg.get("sevk","") else None
    kat_c    = find_col(df, "ANA KATEGORI")
    alt_c    = find_col(df, "ALT KATEGORI")
    aile_c   = find_col(df, "AYAKKABI AILE")
    aalt_c   = find_col(df, "AYAKKABI ALT AILE")
    grp_c    = find_col(df, "SIPARIS GRUBU")
    url_c    = find_col(df, "URL")
    cin_c    = "CINSIYET"

    src_dst = [
        (sevk_c,              "Sevk Tarihi"),
        ("M.R",               "MR OptionKod"),
        (grp_c,               "Siparis Grubu"),
        (cin_c,               "Cinsiyet"),
        (kat_c,               "Ana Kategori"),
        (alt_c,               "Alt Kategori"),
        (aile_c,              "Ayakkabi Aile"),
        (aalt_c,              "Ayakkabi Alt Aile"),
        ("Depo Durumu",       "Depo Durumu"),
        ("Gidis Durumu",      "Gidis Durumu"),
        ("Gidis Yuzde",       "Gidis Yuzde"),
        ("_SIP",              "Toplam Siparis"),
        ("_STOK",             f"{kanal} Kanal Stok"),
        ("_DEPO",             "Depo Stok"),
        ("_WMS",              "WMS Stok"),
        (url_c,               "URL"),
    ]
    mevcut = {s: d for s, d in src_dst if s in sub.columns}
    return sub[list(mevcut.keys())].rename(columns=mevcut)


def excel_gitmeyen(df, cikti_dir: Path):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    dosya = cikti_dir / f"lacoste_gitmeyen_{datetime.now().strftime('%Y%m%d')}.xlsx"
    log(f"Gitmeyen Excel: {dosya.name}", ">>")

    depo_renkler = {
        "MAGAZAYA GITMIS SATIS YOK": "FFD700",
        "DEPODA VAR SEVK BEKLIYOR":  "FFA500",
        "WMSTE VAR":                  "4169E1",
        "DEPOYA GELMEMIS":            "FF69B4",
    }
    gidis_renkler = {
        "GITMEMIS":   ("FF4444","FFFFFF"),
        "KRITIK":     ("FF8C00","FFFFFF"),
        "AZ":         ("FFD700","000000"),
        "IYI":        ("90EE90","000000"),
        "TAMAMLANDI": ("228B22","FFFFFF"),
    }

    with pd.ExcelWriter(dosya, engine="openpyxl") as writer:
        ozet_rows = []
        for kanal in KANALLAR:
            sub = gitmeyen_df(df, kanal)
            if "Depo Durumu" in sub.columns:
                for dd, cnt in sub["Depo Durumu"].value_counts().items():
                    ozet_rows.append({"Kanal": kanal, "Depo Durumu": dd, "Adet": cnt})
            else:
                ozet_rows.append({"Kanal": kanal, "Adet": len(sub)})
        pd.DataFrame(ozet_rows).to_excel(writer, sheet_name="OZET", index=False)

        for kanal in KANALLAR:
            sub = gitmeyen_df(df, kanal)
            if sub.empty:
                pd.DataFrame([{"Bilgi": f"{kanal} icin gitmeyen urun yok"}]).to_excel(
                    writer, sheet_name=kanal, index=False)
            else:
                sub.to_excel(writer, sheet_name=kanal, index=False)

    wb = load_workbook(dosya)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1a1a1a")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if sheet_name == "OZET":
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 4, 40)
            continue

        headers = [cell.value for cell in ws[1]]
        depo_idx  = headers.index("Depo Durumu")  + 1 if "Depo Durumu"  in headers else None
        gidis_idx = headers.index("Gidis Durumu") + 1 if "Gidis Durumu" in headers else None
        pct_idx   = headers.index("Gidis Yuzde")  + 1 if "Gidis Yuzde"  in headers else None

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if depo_idx:
                v = row[depo_idx-1].value
                if v in depo_renkler:
                    row[depo_idx-1].fill = PatternFill("solid", fgColor=depo_renkler[v])
                    row[depo_idx-1].font = Font(bold=True)

            if gidis_idx:
                v = row[gidis_idx-1].value
                if v in gidis_renkler:
                    bg, fg = gidis_renkler[v]
                    row[gidis_idx-1].fill = PatternFill("solid", fgColor=bg)
                    row[gidis_idx-1].font = Font(bold=True, color=fg)

            if pct_idx:
                v = row[pct_idx-1].value
                try:
                    pct = float(v)
                    if pct == 0:     clr = "FF4444"
                    elif pct < 25:   clr = "FF8C00"
                    elif pct < 50:   clr = "FFD700"
                    elif pct < 75:   clr = "90EE90"
                    else:            clr = "228B22"
                    row[pct_idx-1].fill = PatternFill("solid", fgColor=clr)
                    row[pct_idx-1].font = Font(bold=True)
                except (TypeError, ValueError):
                    pass

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 4, 40)

    wb.save(dosya)
    log(f"Kaydedildi: {dosya}", "OK")
    return dosya


JS_DD_FUNCS = """
function toggleDD(fullKey,type){
  var sep=fullKey.indexOf(":");
  var pfx=sep>=0?fullKey.substring(0,sep):fullKey;
  var k=sep>=0?fullKey.substring(sep+1):fullKey;
  var btnId=pfx+"-ddb-"+type+"-"+k;
  var menuId=pfx+"-ddm-"+type+"-"+k;
  var btn=document.getElementById(btnId);
  var dd=document.getElementById(menuId);
  if(!dd||!btn)return;
  var isOpen=dd.classList.contains("open");
  document.querySelectorAll(".dd-menu.open").forEach(function(m){m.classList.remove("open");});
  document.querySelectorAll(".dd-btn.active").forEach(function(b){b.classList.remove("active");});
  if(!isOpen){
    var rect=btn.getBoundingClientRect();
    dd.style.top=(rect.bottom+4)+"px";
    dd.style.left=Math.min(rect.left,window.innerWidth-210)+"px";
    dd.style.maxWidth=(window.innerWidth-20)+"px";
    dd.classList.add("open");
    btn.classList.add("active");
    if(type==="aile"){
      var data=window[pfx+"_URUNLER_"+k]||[];
      var aileler=[...new Set(data.map(function(p){return p.aile;}).filter(function(a){return a&&a!=="0"&&a!=="-"&&a!=="";}) )].sort();
      var list=document.getElementById(pfx+"-aile-list-"+k);
      if(list&&!list.children.length){
        list.innerHTML=aileler.map(function(a){
          return '<label class="dd-item"><input type="checkbox" value="'+a+'"> '+a+'</label>';
        }).join("");
        list.querySelectorAll("input").forEach(function(inp){inp.addEventListener("change",function(){aileChange(fullKey);});});
      }
    }
  }
}
function ddChange(fullKey,type,inp){
  var sep=fullKey.indexOf(":");
  var pfx=sep>=0?fullKey.substring(0,sep):fullKey;
  var k=sep>=0?fullKey.substring(sep+1):fullKey;
  if(!STATE[fullKey])STATE[fullKey]={f:"all",kat:"all",aile:[],s:"top_ytd",d:"desc"};
  if(type==="cin"){
    STATE[fullKey].f=inp.value;
    var lbl=inp.value==="all"?"Tumu":{MEN:"Erkek",WOMEN:"Kadin",CHILDREN:"Cocuk",ACCESSORIES:"Aksesuar"}[inp.value]||inp.value;
    var valEl=document.getElementById(pfx+"-ddv-cin-"+k);
    if(valEl)valEl.textContent=lbl;
    var b=document.getElementById(pfx+"-ddb-cin-"+k);
    if(b){inp.value==="all"?b.classList.remove("active"):b.classList.add("active");}
  } else if(type==="kat"){
    STATE[fullKey].kat=inp.value;
    var valEl=document.getElementById(pfx+"-ddv-kat-"+k);
    if(valEl)valEl.textContent=inp.value==="all"?"Tumu":inp.value;
    var b=document.getElementById(pfx+"-ddb-kat-"+k);
    if(b){inp.value==="all"?b.classList.remove("active"):b.classList.add("active");}
  }
  document.querySelectorAll(".dd-menu.open").forEach(function(m){m.classList.remove("open");});
  document.querySelectorAll(".dd-btn.active").forEach(function(b){b.classList.remove("active");});
  renderGrid(fullKey);
}

// ── ARAMA FONKSİYONLARI (Excel filtre stili) ────────────────
function onSearch(fullKey, val){
  var parts=fullKey.split(":"); var pfx=parts[0]; var k=parts[1];
  STATE[fullKey]=STATE[fullKey]||{f:"all",kat:"all",aile:[],s:"top_ytd",d:"desc",q:"",codes:[]};
  STATE[fullKey].q = val;
  var clearBtn=document.getElementById(pfx+"-srch-clear-"+k);
  if(clearBtn) clearBtn.style.display = val ? "" : "none";
  showSearchDropdown(fullKey, val);
}

function showSearchDropdown(fullKey, q){
  var parts=fullKey.split(":"); var pfx=parts[0]; var k=parts[1];
  var old=document.getElementById(pfx+"-srch-dd-"+k);
  if(old) old.parentNode.removeChild(old);
  var wrap=document.getElementById(pfx+"-srch-wrap-"+k);
  if(!wrap) return;

  var data=getUrunler(pfx,k);
  var st=STATE[fullKey]||{};
  var qn=(q||"").trim().toLowerCase();

  // Eşleşen optionları filtrele
  var matches = qn ? data.filter(function(p){
    return p.code.toLowerCase().indexOf(qn)>=0 ||
           p.ust_kat.toLowerCase().indexOf(qn)>=0 ||
           p.aile.toLowerCase().indexOf(qn)>=0 ||
           p.story.toLowerCase().indexOf(qn)>=0;
  }) : data.slice(0,50);

  if(!matches.length && qn){ renderGrid(fullKey); return; }

  var dd=document.createElement("div");
  dd.id=pfx+"-srch-dd-"+k;
  dd.className="srch-results";

  // Tümünü Seç satırı
  var selAll=document.createElement("div");
  selAll.className="srch-item srch-selall";
  var codes=(st.codes||[]);
  var matchCodes=matches.map(function(p){return p.code;});
  var allSel=matchCodes.length>0&&matchCodes.every(function(c){return codes.indexOf(c)>=0;});
  selAll.innerHTML='<span class="si-chk-box">'+(allSel?'&#9745;':'&#9744;')+'</span>'
    +'<span style="font-weight:700;font-size:10px">Tümünü Seç</span>'
    +'<span class="si-cnt">'+matches.length+' sonuç</span>';
  selAll.onclick=function(e){
    e.stopPropagation();
    if(allSel){
      // Hepsini kaldır
      matchCodes.forEach(function(c){
        var i=STATE[fullKey].codes.indexOf(c);
        if(i>=0) STATE[fullKey].codes.splice(i,1);
      });
    }else{
      // Hepsini ekle
      matchCodes.forEach(function(c){
        if(STATE[fullKey].codes.indexOf(c)<0) STATE[fullKey].codes.push(c);
      });
    }
    updateChips(fullKey);
    showSearchDropdown(fullKey, q);
    renderGrid(fullKey);
  };
  dd.appendChild(selAll);

  // Ayraç
  var div=document.createElement("div");
  div.style.cssText="height:1px;background:var(--border);margin:0 8px";
  dd.appendChild(div);

  // Her option
  matches.forEach(function(p){
    var isSel=(STATE[fullKey].codes||[]).indexOf(p.code)>=0;
    var item=document.createElement("div");
    item.className="srch-item"+(isSel?" sel":"");
    item.innerHTML='<span class="si-chk-box">'+(isSel?'&#9745;':'&#9744;')+'</span>'
      +'<span class="si-code">'+p.code+'</span>'
      +'<span class="si-info">'+p.ust_kat+' · '+p.gender+(p.psf>0?' · '+p.psf.toLocaleString("tr")+'&#8378;':'')+'</span>';
    item.onclick=function(e){
      e.stopPropagation();
      toggleCodeSel(fullKey, p.code);
      showSearchDropdown(fullKey, q);
    };
    dd.appendChild(item);
  });

  // Eşleşme sayısı footer
  if(matches.length>=50&&!qn){
    var footer=document.createElement("div");
    footer.style.cssText="padding:5px 10px;font-size:9px;color:var(--text3);text-align:center";
    footer.textContent="İlk 50 gösteriliyor — aramak için yazmaya başlayın";
    dd.appendChild(footer);
  }

  wrap.appendChild(dd);
}

function toggleCodeSel(fullKey, code){
  STATE[fullKey]=STATE[fullKey]||{f:"all",kat:"all",aile:[],s:"top_ytd",d:"desc",q:"",codes:[]};
  var codes=STATE[fullKey].codes||[];
  var idx=codes.indexOf(code);
  if(idx>=0){ codes.splice(idx,1); }
  else{ codes.push(code); }
  STATE[fullKey].codes=codes;
  updateChips(fullKey);
  renderGrid(fullKey);
}

function updateChips(fullKey){
  var parts=fullKey.split(":"); var pfx=parts[0]; var k=parts[1];
  var codes=(STATE[fullKey]||{}).codes||[];
  var container=document.getElementById(pfx+"-chips-"+k);
  if(!container) return;
  container.innerHTML="";
  if(codes.length>0){
    // Özet chip: "L1212 (8 seçili)" gibi
    var grp={};
    codes.forEach(function(c){
      var prefix=c.split(".")[0];
      if(!grp[prefix]) grp[prefix]=0;
      grp[prefix]++;
    });
    Object.keys(grp).forEach(function(prefix){
      var chip=document.createElement("div");
      chip.className="sel-chip";
      chip.title="Bu gruptaki seçimleri kaldır";
      chip.innerHTML='<span>'+prefix+(grp[prefix]>1?' ('+grp[prefix]+')':'')+'</span><span class="ch-x">&#10005;</span>';
      chip.onclick=function(){
        // Bu prefix'e ait tüm kodları kaldır
        STATE[fullKey].codes=STATE[fullKey].codes.filter(function(c){return c.split(".")[0]!==prefix;});
        updateChips(fullKey);
        var inp=document.getElementById(pfx+"-srch-"+k);
        if(inp) showSearchDropdown(fullKey, inp.value);
        renderGrid(fullKey);
      };
      container.appendChild(chip);
    });
    // Tümünü Temizle
    var clearAll=document.createElement("div");
    clearAll.className="sel-chip sel-chip-clear";
    clearAll.innerHTML='<span>Tümünü Temizle</span><span class="ch-x">&#10005;</span>';
    clearAll.onclick=function(){ clearSearch(fullKey); };
    container.appendChild(clearAll);
  }
}

function clearSearch(fullKey){
  var parts=fullKey.split(":"); var pfx=parts[0]; var k=parts[1];
  if(!STATE[fullKey]) return;
  STATE[fullKey].q="";
  STATE[fullKey].codes=[];
  var inp=document.getElementById(pfx+"-srch-"+k);
  if(inp) inp.value="";
  var clearBtn=document.getElementById(pfx+"-srch-clear-"+k);
  if(clearBtn) clearBtn.style.display="none";
  var dd=document.getElementById(pfx+"-srch-dd-"+k);
  if(dd) dd.parentNode.removeChild(dd);
  updateChips(fullKey);
  renderGrid(fullKey);
}

// Enter ile Tümünü Seç
document.addEventListener("keydown",function(e){
  if(e.key!=="Enter") return;
  var inp=document.activeElement;
  if(!inp||!inp.classList.contains("srch-inp")) return;
  // Hangi input?
  var id=inp.id; // "AW-srch-SSTEP" formatı
  var m=id.match(/^(.+)-srch-(.+)$/);
  if(!m) return;
  var pfx=m[1], k=m[2];
  var fullKey=pfx+":"+k;
  var q=inp.value.trim();
  if(!q) return;
  var data=getUrunler(pfx,k);
  var qn=q.toLowerCase();
  var matches=data.filter(function(p){
    return p.code.toLowerCase().indexOf(qn)>=0||
           p.ust_kat.toLowerCase().indexOf(qn)>=0||
           p.aile.toLowerCase().indexOf(qn)>=0||
           p.story.toLowerCase().indexOf(qn)>=0;
  });
  // Hepsini seç
  matches.forEach(function(p){
    if((STATE[fullKey].codes||[]).indexOf(p.code)<0)
      STATE[fullKey].codes.push(p.code);
  });
  updateChips(fullKey);
  showSearchDropdown(fullKey, q);
  renderGrid(fullKey);
});

// ── AİLE ARAMA FONKSİYONLARI ─────────────────────────
function onAileSearch(fullKey, val) {
  var parts=fullKey.split(":"); var pfx=parts[0]; var k=parts[1];
  STATE[fullKey] = STATE[fullKey] || {f:"all",kat:"all",aile:[],aile_codes:[],s:"top_ytd",d:"desc",q:"",codes:[]};
  STATE[fullKey].aile_q = val;
  var clearBtn = document.getElementById(pfx+"-aile-srch-clear-"+k);
  if(clearBtn) clearBtn.style.display = val ? "" : "none";
  showAileDropdown(fullKey, val);
}

function showAileDropdown(fullKey, q) {
  var parts=fullKey.split(":"); var pfx=parts[0]; var k=parts[1];
  var old = document.getElementById(pfx+"-aile-srch-dd-"+k);
  if(old) old.parentNode.removeChild(old);
  var wrap = document.getElementById(pfx+"-aile-srch-wrap-"+k);
  if(!wrap) return;

  // Tüm aile değerlerini topla
  var data = getUrunler(pfx,k);
  var aileMap = {};
  data.forEach(function(p) {
    if(p.aile && p.aile !== "-") aileMap[p.aile] = (aileMap[p.aile]||0) + 1;
  });
  var allAiles = Object.keys(aileMap).sort();
  var qn = (q||"").trim().toLowerCase();
  var matches = qn ? allAiles.filter(function(a){ return a.toLowerCase().indexOf(qn) >= 0; }) : allAiles;
  if(!matches.length) return;

  var st = STATE[fullKey] || {};
  var selCodes = st.aile_codes || [];

  var dd = document.createElement("div");
  dd.id = pfx+"-aile-srch-dd-"+k;
  dd.className = "srch-results";

  // Tümünü Seç
  var allSel = matches.length > 0 && matches.every(function(a){ return selCodes.indexOf(a) >= 0; });
  var selAll = document.createElement("div");
  selAll.className = "srch-item srch-selall";
  selAll.innerHTML = '<span class="si-chk-box">'+(allSel?'&#9745;':'&#9744;')+'</span>'
    +'<span style="font-weight:700;font-size:10px">Tümünü Seç</span>'
    +'<span class="si-cnt">'+matches.length+' aile</span>';
  selAll.onclick = function(e) {
    e.stopPropagation();
    STATE[fullKey].aile_codes = STATE[fullKey].aile_codes || [];
    if(allSel) {
      matches.forEach(function(a) {
        var i = STATE[fullKey].aile_codes.indexOf(a);
        if(i >= 0) STATE[fullKey].aile_codes.splice(i,1);
      });
    } else {
      matches.forEach(function(a) {
        if(STATE[fullKey].aile_codes.indexOf(a) < 0) STATE[fullKey].aile_codes.push(a);
      });
    }
    updateAileChips(fullKey);
    showAileDropdown(fullKey, q);
    renderGrid(fullKey);
  };
  dd.appendChild(selAll);

  var div = document.createElement("div");
  div.style.cssText = "height:1px;background:var(--border);margin:0 8px";
  dd.appendChild(div);

  matches.forEach(function(aile) {
    var cnt = aileMap[aile] || 0;
    var isSel = selCodes.indexOf(aile) >= 0;
    var item = document.createElement("div");
    item.className = "srch-item" + (isSel ? " sel" : "");
    item.innerHTML = '<span class="si-chk-box">'+(isSel?'&#9745;':'&#9744;')+'</span>'
      +'<span class="si-code">'+aile+'</span>'
      +'<span class="si-info">'+cnt+' ürün</span>';
    item.onclick = function(e) {
      e.stopPropagation();
      toggleAileSel(fullKey, aile);
      showAileDropdown(fullKey, q);
    };
    dd.appendChild(item);
  });

  wrap.appendChild(dd);
}

function toggleAileSel(fullKey, aile) {
  STATE[fullKey] = STATE[fullKey] || {f:"all",kat:"all",aile:[],aile_codes:[],s:"top_ytd",d:"desc",q:"",codes:[]};
  var codes = STATE[fullKey].aile_codes || [];
  var idx = codes.indexOf(aile);
  if(idx >= 0) codes.splice(idx,1);
  else codes.push(aile);
  STATE[fullKey].aile_codes = codes;
  updateAileChips(fullKey);
  renderGrid(fullKey);
}

function updateAileChips(fullKey) {
  var parts=fullKey.split(":"); var pfx=parts[0]; var k=parts[1];
  var codes = (STATE[fullKey]||{}).aile_codes || [];
  var container = document.getElementById(pfx+"-aile-chips-"+k);
  if(!container) return;
  container.innerHTML = "";
  codes.forEach(function(aile) {
    var chip = document.createElement("div");
    chip.className = "sel-chip";
    chip.style.cssText = "font-size:8px;padding:2px 6px 2px 8px";
    chip.innerHTML = '<span>'+aile+'</span><span class="ch-x">&#10005;</span>';
    chip.onclick = function() { toggleAileSel(fullKey, aile); showAileDropdown(fullKey, document.getElementById(pfx+"-aile-srch-"+k)?document.getElementById(pfx+"-aile-srch-"+k).value:""); };
    container.appendChild(chip);
  });
  if(codes.length > 0) {
    var cl = document.createElement("div");
    cl.className = "sel-chip sel-chip-clear";
    cl.style.cssText = "font-size:8px;padding:2px 6px";
    cl.innerHTML = '<span>Temizle</span><span class="ch-x">&#10005;</span>';
    cl.onclick = function() { clearAileSearch(fullKey); };
    container.appendChild(cl);
  }
}

function clearAileSearch(fullKey) {
  var parts=fullKey.split(":"); var pfx=parts[0]; var k=parts[1];
  if(!STATE[fullKey]) return;
  STATE[fullKey].aile_codes = [];
  STATE[fullKey].aile_q = "";
  var inp = document.getElementById(pfx+"-aile-srch-"+k);
  if(inp) inp.value = "";
  var clearBtn = document.getElementById(pfx+"-aile-srch-clear-"+k);
  if(clearBtn) clearBtn.style.display = "none";
  var dd = document.getElementById(pfx+"-aile-srch-dd-"+k);
  if(dd) dd.parentNode.removeChild(dd);
  updateAileChips(fullKey);
  renderGrid(fullKey);
}

// Enter ile tümünü seç (aile)
document.addEventListener("keydown", function(e) {
  if(e.key !== "Enter") return;
  var inp = document.activeElement;
  if(!inp || inp.id.indexOf("-aile-srch-") < 0) return;
  var m = inp.id.match(/^(.+)-aile-srch-(.+)$/);
  if(!m) return;
  var pfx=m[1], k=m[2], fullKey=pfx+":"+k;
  var q = inp.value.trim();
  var data = getUrunler(pfx,k);
  var aileMap = {};
  data.forEach(function(p){ if(p.aile&&p.aile!=="-") aileMap[p.aile]=true; });
  var qn = q.toLowerCase();
  Object.keys(aileMap).filter(function(a){ return !qn||a.toLowerCase().indexOf(qn)>=0; }).forEach(function(a) {
    STATE[fullKey].aile_codes = STATE[fullKey].aile_codes || [];
    if(STATE[fullKey].aile_codes.indexOf(a) < 0) STATE[fullKey].aile_codes.push(a);
  });
  updateAileChips(fullKey);
  showAileDropdown(fullKey, q);
  renderGrid(fullKey);
});

// Dışarı tıklanınca tüm dropdown'ları kapat
document.addEventListener("click",function(){
  document.querySelectorAll(".srch-results").forEach(function(d){
    if(d.parentNode) d.parentNode.removeChild(d);
  });
});
"""



def uret_html_cift(df_aw, df_ss, hafta_str, aw_adi, ss_adi):
    """
    26AW ve 26SS verilerini birlestirir.
    Ortada sezon toggle butonu - tiklayinca ilgili veriler gelir.
    Her sezonun JS verileri farkli prefix ile tutulur: AW_URUNLER_SSTEP, SS_URUNLER_SSTEP
    """
    import json

    J = lambda o: json.dumps(o, ensure_ascii=True)

    sezonlar = {
        "26AW": (df_aw, aw_adi),
        "26SS": (df_ss, ss_adi),
    }

    # Her sezon icin veri hazirla
    sezon_data = {}
    for sezon, (df, excel_adi) in sezonlar.items():
        kanal_data = {}
        for kanal in KANALLAR:
            urunler = urun_listesi(df, kanal)
            analiz  = str_analiz(df, kanal)
            gitm    = gitmeyen_df(df, kanal)
            kanal_data[kanal] = {
                "urunler":  urunler,
                "analiz":   analiz,
                "git_cnt":  len(gitm),
            }
        hoss = hoss_analiz(df)
        sezon_data[sezon] = {
            "kanal_data": kanal_data,
            "hoss":       hoss,
            "total":      len(df),
            "excel_adi":  excel_adi,
        }

    # JS veri bloklari - her sezon icin prefix'li
    def urun_js_blok(rows, prefix):
        items = []
        for p in rows:
            items.append(
                "{" + f"code:{J(p['code'])},gender:{J(p['gender'])},ana_kat:{J(p['ana_kat'])},"
                f"ust_kat:{J(p['ust_kat'])},alt_kat:{J(p['alt_kat'])},aile:{J(p['aile'])},"
                f"alt_aile:{J(p['alt_aile'])},sip_grp:{J(p['sip_grp'])},"
                f"sip_h:{p['sip_h']},sip_d:{p['sip_d']},"
                f"mag_ytd:{p['mag_ytd']},onl_ytd:{p['onl_ytd']},top_ytd:{p['top_ytd']},"
                f"mag_sh:{p['mag_sh']},onl_sh:{p['onl_sh']},top_sh:{p['top_sh']},"
                f"mag_mtd:{p.get('mag_mtd',0)},onl_mtd:{p.get('onl_mtd',0)},top_mtd:{p.get('top_mtd',0)},"
                f"stok:{p['stok']},ros:{p['ros']},str_h:{p['str_h']},str_d:{p['str_d']},"
                f"story:{J(p.get('story',''))},is_hoss:{'true' if p.get('is_hoss') else 'false'},"
                f"psf:{p.get('psf',0)},"
                f"wms:{p.get('wms',0)},"
                f"depo:{p.get('depo',0)},"
                f"img:{J(p['img'])}" + "}"
            )
        return "[\n" + ",\n".join(items) + "\n]"

    js_data_bloklar = []
    for sezon, sd in sezon_data.items():
        pfx = sezon.replace("26","")    # 26AW -> AW, 26SS -> SS
        for kanal in KANALLAR:
            rows = sd["kanal_data"][kanal]["urunler"]
            vname_u = f'{pfx}_URUNLER_{kanal}'
            vname_a = f'{pfx}_ANALIZ_{kanal}'
            js_data_bloklar.append(f'var {vname_u}={urun_js_blok(rows, pfx)};window["{vname_u}"]={vname_u};')
            js_data_bloklar.append(f'var {vname_a}={J(sd["kanal_data"][kanal]["analiz"])};window["{vname_a}"]={vname_a};')
        hoss = sd["hoss"]
        for vn, val in [
            (f'{pfx}_HOSS_URUNLER', _hoss_urun_js(J, hoss.get("urunler",[]))),
            (f'{pfx}_HOSS_KAT',     J(hoss.get("ozet_kat", []))),
            (f'{pfx}_HOSS_UST',     J(hoss.get("ozet_ust", []))),
            (f'{pfx}_HOSS_CIN',     J(hoss.get("ozet_cin", []))),
        ]:
            js_data_bloklar.append(f'var {vn}={val};window["{vn}"]={vn};')

    js_data = "\n".join(js_data_bloklar)

    kmeta_aw = J({k: {"renk": KANALLAR[k]["renk"], "git_cnt": sezon_data["26AW"]["kanal_data"][k]["git_cnt"]} for k in KANALLAR if len(sezon_data["26AW"]["kanal_data"][k]["urunler"]) > 0})
    kmeta_ss = J({k: {"renk": KANALLAR[k]["renk"], "git_cnt": sezon_data["26SS"]["kanal_data"][k]["git_cnt"]} for k in KANALLAR if len(sezon_data["26SS"]["kanal_data"][k]["urunler"]) > 0})

    # CSS'i uret_html'den al (tek seferlik, sadece style bloğu)
    _tmp_html = uret_html(df_aw, "tmp", "tmp.xlsx")
    css_blok = _tmp_html[_tmp_html.find("<style>"):_tmp_html.find("</style>")+8]
    del _tmp_html

    # Kanal panel HTML'leri üret
    def kanal_paneller(sd, sezon_pfx):
        paneller = ""
        hoss_count = sd["hoss"].get("count", 0)
        hoss_js_kat  = J(sd["hoss"].get("ozet_kat", []))
        hoss_js_ust  = J(sd["hoss"].get("ozet_ust", []))
        hoss_js_cin  = J(sd["hoss"].get("ozet_cin", []))
        for k, cfg in KANALLAR.items():
            if len(sd["kanal_data"][k]["urunler"]) == 0:
                continue  # Ürün yok, sekme/panel oluşturma
            git = sd["kanal_data"][k]["git_cnt"]
            paneller += f"""
<div class="panel" id="{sezon_pfx}-panel-{k}">
  <div class="kpi-bar" style="border-bottom:3px solid {cfg['renk']}">
    <div class="kpi"><div class="kpi-l">Sip. Devir Haric</div><div class="kpi-v" id="{sezon_pfx}-kv1-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Sip. Devir Dahil</div><div class="kpi-v" id="{sezon_pfx}-kv2-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Mag Satis YTD</div><div class="kpi-v" id="{sezon_pfx}-kv3-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Online Satis YTD</div><div class="kpi-v" id="{sezon_pfx}-kv4-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Toplam Satis YTD</div><div class="kpi-v" id="{sezon_pfx}-kv5-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">STR D.Haric</div><div class="kpi-v" style="color:{cfg['renk']}" id="{sezon_pfx}-kv6-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">STR D.Dahil</div><div class="kpi-v" style="color:{cfg['renk']}" id="{sezon_pfx}-kv7-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Kanal Stok</div><div class="kpi-v" id="{sezon_pfx}-kv8-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Gitmeyen</div><div class="kpi-v">{git} <button class="dlb" onclick="indir('{k}')">Excel</button></div></div>
  </div>
  <div class="sub-tabs">
    <button class="stab active" onclick="switchSub('{sezon_pfx}-{k}','u',this)">Tum Urunler</button>
    <button class="stab" onclick="switchSub('{sezon_pfx}-{k}','s',this)">STR Analizi</button>
  </div>
  <div class="sub-panel active" id="sp-{sezon_pfx}-{k}-u">
    <div class="toolbar2">
      <div class="search-bar">
        <div class="srch-wrap srch-wrap-rel" id="{sezon_pfx}-srch-wrap-{k}">
          <span class="srch-icon">&#128269;</span>
          <input class="srch-inp" id="{sezon_pfx}-srch-{k}" type="text" placeholder="Option ara... (Enter = Tümünü Seç)"
            oninput="onSearch('{sezon_pfx}:{k}',this.value)"
            onclick="event.stopPropagation();showSearchDropdown('{sezon_pfx}:{k}',this.value)"
            onfocus="showSearchDropdown('{sezon_pfx}:{k}',this.value)">
          <button class="srch-clear" id="{sezon_pfx}-srch-clear-{k}" onclick="clearSearch('{sezon_pfx}:{k}')" style="display:none">&#10005;</button>
        </div>
        <div class="sel-chips" id="{sezon_pfx}-chips-{k}"></div>
      </div>
      <div class="tb-left">
        <div class="dd-wrap" id="{sezon_pfx}-ddw-cin-{k}">
          <button class="dd-btn" id="{sezon_pfx}-ddb-cin-{k}" onclick="toggleDD('{sezon_pfx}:{k}','cin')">Cinsiyet <span class="dd-val" id="{sezon_pfx}-ddv-cin-{k}">Tumu</span> &#9662;</button>
          <div class="dd-menu" id="{sezon_pfx}-ddm-cin-{k}">
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-cin-{k}" value="all" checked onchange="ddChange('{sezon_pfx}:{k}','cin',this)"> Tumu</label>
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-cin-{k}" value="MEN" onchange="ddChange('{sezon_pfx}:{k}','cin',this)"> Erkek</label>
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-cin-{k}" value="WOMEN" onchange="ddChange('{sezon_pfx}:{k}','cin',this)"> Kadin</label>
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-cin-{k}" value="CHILDREN" onchange="ddChange('{sezon_pfx}:{k}','cin',this)"> Cocuk</label>
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-cin-{k}" value="ACCESSORIES" onchange="ddChange('{sezon_pfx}:{k}','cin',this)"> Aksesuar</label>
          </div>
        </div>
        <div class="dd-wrap" id="{sezon_pfx}-ddw-kat-{k}">
          <button class="dd-btn" id="{sezon_pfx}-ddb-kat-{k}" onclick="toggleDD('{sezon_pfx}:{k}','kat')">Kategori <span class="dd-val" id="{sezon_pfx}-ddv-kat-{k}">Tumu</span> &#9662;</button>
          <div class="dd-menu" id="{sezon_pfx}-ddm-kat-{k}">
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-kat-{k}" value="all" checked onchange="ddChange('{sezon_pfx}:{k}','kat',this)"> Tumu</label>
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-kat-{k}" value="FOOTWEAR" onchange="ddChange('{sezon_pfx}:{k}','kat',this)"> Footwear</label>
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-kat-{k}" value="TEXTILE" onchange="ddChange('{sezon_pfx}:{k}','kat',this)"> Textile</label>
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-kat-{k}" value="TEXTILE ACCESSORIES" onchange="ddChange('{sezon_pfx}:{k}','kat',this)"> Textile Acc.</label>
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-kat-{k}" value="ACCESSORIES" onchange="ddChange('{sezon_pfx}:{k}','kat',this)"> Accessories</label>
            <label class="dd-item"><input type="radio" name="{sezon_pfx}-kat-{k}" value="UNDERWEAR" onchange="ddChange('{sezon_pfx}:{k}','kat',this)"> Underwear</label>
          </div>
        </div>
        <div class="dd-wrap" id="{sezon_pfx}-ddw-aile-{k}">
          <div class="srch-wrap srch-wrap-rel" id="{sezon_pfx}-aile-srch-wrap-{k}" style="min-width:130px">
            <span class="srch-icon" style="font-size:11px">&#128270;</span>
            <input class="srch-inp" id="{sezon_pfx}-aile-srch-{k}" type="text" placeholder="Aile ara..."
              oninput="onAileSearch('{sezon_pfx}:{k}',this.value)"
              onclick="event.stopPropagation();showAileDropdown('{sezon_pfx}:{k}',this.value)"
              onfocus="showAileDropdown('{sezon_pfx}:{k}',this.value)">
            <button class="srch-clear" id="{sezon_pfx}-aile-srch-clear-{k}"
              onclick="clearAileSearch('{sezon_pfx}:{k}')" style="display:none">&#10005;</button>
          </div>
          <div class="sel-chips" id="{sezon_pfx}-aile-chips-{k}" style="margin-top:3px"></div>
        </div>
      </div>
      <div class="tb-right">
        <span class="tlbl2">Sirala</span>
        <div id="{sezon_pfx}-sg-{k}" class="sort-group">
          <button class="sbtn as" data-s="top_ytd" data-d="desc">Tum Satislar</button>
          <button class="sbtn" data-s="mag_ytd" data-d="desc">Magaza YTD</button>
          <button class="sbtn" data-s="onl_ytd" data-d="desc">Online YTD</button>
          <button class="sbtn" data-s="mag_mtd" data-d="desc">MTD Mag</button>
          <button class="sbtn" data-s="onl_mtd" data-d="desc">MTD Online</button>
          <button class="sbtn" data-s="top_sh" data-d="desc">Son Hafta Mag</button>
          <button class="sbtn" data-s="onl_sh" data-d="desc">Son Hafta Online</button>
          <button class="sbtn" data-s="stok" data-d="desc">Magaza Stok</button>
          <button class="sbtn" data-s="str_h" data-d="desc">STR</button>
          <button class="sbtn" data-s="ros" data-d="desc">ROS</button>
          <button class="sbtn" data-s="top_ytd" data-d="asc">En Az Satanlar</button>
        </div>
      </div>
    </div>
    <div class="gw">
      <div class="shdr">
        <span id="{sezon_pfx}-lbl-{k}">Toplam Satisa Gore</span>
        <span id="{sezon_pfx}-cnt-{k}">-</span>
      </div>
      <div class="tot-bar" id="{sezon_pfx}-tot-{k}"></div>
      <button class="zero-btn" id="{sezon_pfx}-zerobtn-{k}" onclick="toggleZero('{sezon_pfx}','{k}')" style="display:none">
        <span class="zarr">&#9658;</span> Sifir Satislar
      </button>
      <div class="grid" id="{sezon_pfx}-grid-{k}"></div>
    </div>
  </div>
  <div class="sub-panel" id="sp-{sezon_pfx}-{k}-s">
    <div class="str-nav">
      <button class="snav act" onclick="switchStrView('{sezon_pfx}:{k}','kat',this)">Ana Kategori</button>
      <button class="snav" onclick="switchStrView('{sezon_pfx}:{k}','cin',this)">Cinsiyet</button>
      <button class="snav" onclick="switchStrView('{sezon_pfx}:{k}','aile',this)">Ayakkabi Aile</button>
    </div>
    <div class="str-wrap">
      <div class="str-sub active" id="ss-{sezon_pfx}-{k}-kat">
        <div class="str-section"><div class="str-title" style="border-color:{cfg['renk']}">Ana Kategori Bazli STR</div><div class="stw" id="{sezon_pfx}-st-{k}"></div></div>
      </div>
      <div class="str-sub" id="ss-{sezon_pfx}-{k}-cin">
        <div class="str-section"><div class="str-title" style="border-color:{cfg['renk']}">Cinsiyet Bazli STR</div><div class="stw" id="{sezon_pfx}-st-cin-{k}"></div></div>
      </div>
      <div class="str-sub" id="ss-{sezon_pfx}-{k}-aile">
        <div class="str-section"><div class="str-title" style="border-color:{cfg['renk']}">Ayakkabi Aile Bazli STR</div><div class="stw" id="{sezon_pfx}-st-aile-{k}"></div></div>
      </div>
    </div>
  </div>
</div>"""
        # HOSS paneli
        paneller += f"""
<div class="panel" id="{sezon_pfx}-panel-HOSS">
  <div class="hoss-kpi">
    <div class="kpi"><div class="kpi-l">Toplam Urun</div><div class="kpi-v">{hoss_count}</div></div>
    <div class="kdiv" style="background:rgba(196,122,58,.3)"></div>
    <div class="kpi"><div class="kpi-l">Koleksiyon</div><div class="kpi-v" style="font-size:14px">HOSS x Lacoste Kapsul</div></div>
  </div>
  <div class="sub-tabs">
    <button class="stab active" onclick="switchSub('{sezon_pfx}-HOSS','u',this)">Tum Urunler</button>
    <button class="stab" onclick="switchSub('{sezon_pfx}-HOSS','s',this)">STR Analizi</button>
    <button class="stab" onclick="switchSub('{sezon_pfx}-HOSS','o',this)">Ozet</button>
  </div>
  <div class="sub-panel active" id="sp-{sezon_pfx}-HOSS-u">
    <div class="toolbar2">
      <div class="search-bar">
        <div class="srch-wrap srch-wrap-rel" id="{sezon_pfx}-srch-wrap-{k}">
          <span class="srch-icon">&#128269;</span>
          <input class="srch-inp" id="{sezon_pfx}-srch-{k}" type="text" placeholder="Option ara... (Enter = Tümünü Seç)"
            oninput="onSearch('{sezon_pfx}:{k}',this.value)"
            onclick="event.stopPropagation();showSearchDropdown('{sezon_pfx}:{k}',this.value)"
            onfocus="showSearchDropdown('{sezon_pfx}:{k}',this.value)">
          <button class="srch-clear" id="{sezon_pfx}-srch-clear-{k}" onclick="clearSearch('{sezon_pfx}:{k}')" style="display:none">&#10005;</button>
        </div>
        <div class="sel-chips" id="{sezon_pfx}-chips-{k}"></div>
      </div>
      <div class="tb-left">
        <div id="{sezon_pfx}-fg-HOSS" style="display:flex;gap:5px">
          <button class="btn af" data-f="all">Tumu</button>
          <button class="btn" data-f="MEN">Erkek</button>
          <button class="btn" data-f="WOMEN">Kadin</button>
          <button class="btn" data-f="ACCESSORIES">Aksesuar</button>
        </div>
      </div>
      <div class="tb-right">
        <span class="tlbl2">Kanal</span>
        <div id="{sezon_pfx}-sg-HOSS" style="display:flex;gap:4px">
          {''.join([f'<button class="sbtn{" as" if i==0 else ""}" data-k="{k}">{k}</button>' for i,k in enumerate(KANALLAR)])}
        </div>
      </div>
    </div>
    <div class="gw"><div class="shdr"><span id="{sezon_pfx}-lbl-HOSS">SSTEP Satisina Gore</span><span id="{sezon_pfx}-cnt-HOSS">-</span></div>
    <div class="grid" id="{sezon_pfx}-grid-HOSS"></div></div>
  </div>
  <div class="sub-panel" id="sp-{sezon_pfx}-HOSS-s">
    <div class="str-wrap">
      <div class="str-section"><div class="str-title" style="border-color:#8B5E3C">Ana Kategori</div><div class="stw" id="{sezon_pfx}-st-HOSS-kat"></div></div>
      <div class="str-section"><div class="str-title" style="border-color:#8B5E3C">Ust Kategori</div><div class="stw" id="{sezon_pfx}-st-HOSS-ust"></div></div>
      <div class="str-section"><div class="str-title" style="border-color:#8B5E3C">Kod Bazli STR</div><div class="stw" id="{sezon_pfx}-st-HOSS-kod"></div></div>
    </div>
  </div>
  <div class="sub-panel" id="sp-{sezon_pfx}-HOSS-o">
    <div class="str-wrap"><div class="str-section"><div class="str-title" style="border-color:#8B5E3C">Cinsiyet Ozet</div><div class="stw" id="{sezon_pfx}-st-HOSS-cin"></div></div></div>
  </div>
</div>"""
        return paneller

    # Her sezon icin tab bar
    def tab_bar(sd, sezon_pfx):
        tabs = "\n".join([
            f'<button class="tab" style="--tk:{KANALLAR[k]["renk"]}" onclick="switchTab(\'{k}\',this,\'{sezon_pfx}\')">{k} <span class="tc">{sd["kanal_data"][k]["git_cnt"]} bekleyen</span></button>'
            for k in KANALLAR
            if len(sd["kanal_data"][k]["urunler"]) > 0
        ])
        if sd["hoss"].get("count", 0) > 0:
            tabs += f'\n<button class="tab hoss-tab" onclick="switchTab(\'HOSS\',this,\'{sezon_pfx}\')">&#x2726; HOSS x Lacoste <span class="tc">{sd["hoss"].get("count",0)} urun</span></button>'
        return f'<div class="tab-bar" id="{sezon_pfx}-tab-bar">{tabs}</div>'

    panels_aw = kanal_paneller(sezon_data["26AW"], "AW")
    panels_ss = kanal_paneller(sezon_data["26SS"], "SS")
    tabs_aw   = tab_bar(sezon_data["26AW"], "AW")
    tabs_ss   = tab_bar(sezon_data["26SS"], "SS")

    sezon_toggle_html = f"""
<div class="sezon-bar">
  <div class="sezon-toggle">
    <button class="sezon-btn active" id="sz-btn-AW" onclick="switchSezon('AW')">&#9670; 26AW</button>
    <button class="sezon-btn" id="sz-btn-SS" onclick="switchSezon('SS')">&#9670; 26SS</button>
  </div>
</div>"""

    header_html = f"""<div class="header">
  <div class="logo-wrap">
    <div class="logo-icon">&#x1F40A;</div>
    <div>
      <div class="brand">Lacoste Sezon Takip Raporu</div>
      <div class="brand-sub">{hafta_str}</div>
    </div>
  </div>
  <div class="hb">{sezon_data['26AW']['total']+sezon_data['26SS']['total']} Option</div>
</div>"""

    footer_html = f'<div class="footer">STR = (Mag+Online) / Siparis x 100 &nbsp;&middot;&nbsp; <b>{hafta_str}</b></div>'

    sezon_css = """<style>
.sezon-bar{background:#1a1008;padding:.45rem 1rem;display:flex;justify-content:center;border-bottom:2px solid var(--gold)}
.sezon-toggle{display:flex;gap:0;background:#2c2010;border-radius:10px;padding:3px;border:1px solid var(--gold2)}
.sezon-btn{padding:7px 22px;font-size:12px;font-weight:700;cursor:pointer;border:none;background:none;color:#8a7a5a;border-radius:8px;font-family:'DM Sans',sans-serif;transition:all .2s;letter-spacing:.5px;white-space:nowrap}
.sezon-btn.active{background:var(--gold);color:var(--dark);box-shadow:0 2px 8px rgba(200,168,75,.3)}
.sezon-btn:hover:not(.active){color:#f0e8d0}
.sezon-wrap{display:none}.sezon-wrap.active{display:block}
@media(max-width:600px){.sezon-btn{padding:6px 14px;font-size:11px}}
</style>"""

    _html = f"""<!DOCTYPE html>
<html lang="tr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1">
<title>Lacoste Sezon Takip &mdash; {hafta_str}</title>
{css_blok}
{sezon_css}
</head>
<body>
{header_html}
{sezon_toggle_html}
<div class="sezon-wrap active" id="sz-wrap-AW">
{tabs_aw}
{panels_aw}
</div>
<div class="sezon-wrap" id="sz-wrap-SS">
{tabs_ss}
{panels_ss}
</div>
{footer_html}
"""

    _html += f"""<script>
""" + JS_DD_FUNCS + f"""
{js_data}
var AW_KMETA={kmeta_aw};
var SS_KMETA={kmeta_ss};
window["AW_KMETA"]=AW_KMETA;
window["SS_KMETA"]=SS_KMETA;
var KMETA=AW_KMETA;
var CUR_PFX="AW";
var STATE={{}};
var HOSS_STATE={{}};
for(var _k in AW_KMETA){{STATE["AW:"+_k]={{f:"all",kat:"all",aile:[],aile_codes:[],s:"top_ytd",d:"desc",q:"",codes:[]}};HOSS_STATE["AW:"+_k]={{f:"all",sort_k:"SSTEP"}};}}
for(var _k in SS_KMETA){{STATE["SS:"+_k]={{f:"all",kat:"all",aile:[],aile_codes:[],s:"top_ytd",d:"desc",q:"",codes:[]}};HOSS_STATE["SS:"+_k]={{f:"all",sort_k:"SSTEP"}};}}
var gL={{MEN:"Erkek",WOMEN:"Kadin",CHILDREN:"Cocuk",ACCESSORIES:"Aksesuar"}};
function sc(v){{return v>=20?"sg1":v>=10?"sg2":"sg3";}}

function getUrunler(pfx,k){{return window[pfx+"_URUNLER_"+k]||[];}}
function getAnaliz(pfx,k){{return window[pfx+"_ANALIZ_"+k]||{{}};}}

function renderGrid(fullKey){{
  var parts=fullKey.split(":");var pfx=parts[0];var k=parts[1];
  var grid=document.getElementById(pfx+"-grid-"+k);if(!grid)return;
  var st=STATE[fullKey]||{{f:"all",kat:"all",aile:[],s:"top_ytd",d:"desc"}};
  var data=getUrunler(pfx,k);
  var vis=[...data];
  if(st.f!=="all")vis=vis.filter(function(p){{return p.gender===st.f;}});
  if(st.kat&&st.kat!=="all")vis=vis.filter(function(p){{return p.ana_kat===st.kat;}});
  if(st.aile&&st.aile.length>0)vis=vis.filter(function(p){{return st.aile.indexOf(p.aile)>=0;}});
  if(st.aile_codes&&st.aile_codes.length>0)vis=vis.filter(function(p){{return st.aile_codes.indexOf(p.aile)>=0;}});
  // Arama filtresi (q)
  if(st.q&&st.q.trim()){{
    var qn=st.q.trim().toLowerCase();
    vis=vis.filter(function(p){{
      return p.code.toLowerCase().indexOf(qn)>=0||
             p.ust_kat.toLowerCase().indexOf(qn)>=0||
             p.aile.toLowerCase().indexOf(qn)>=0||
             p.story.toLowerCase().indexOf(qn)>=0||
             String(p.psf).indexOf(qn)>=0;
    }});
  }}
  // Çoklu option seçimi (codes)
  if(st.codes&&st.codes.length>0){{
    vis=vis.filter(function(p){{return st.codes.indexOf(p.code)>=0;}});
  }}
  var asc=st.d==="asc";
  vis.sort(function(a,b){{return asc?a[st.s]-b[st.s]:b[st.s]-a[st.s];}});
  var KMETA_CUR=window[pfx+"_KMETA"]||KMETA;
  var renk=(KMETA_CUR[k]||{{renk:"#2c2416"}}).renk;
  var maxV=Math.max.apply(null,vis.map(function(p){{return Math.max(0,p[st.s]||0);}}).concat([1]));
  grid.innerHTML="";
  vis.forEach(function(p,i){{
    var isTop=!asc&&i<3;
    var card=document.createElement("div");
    card.className="card"+(isTop?" top":"");
    if(isTop)card.style.borderColor=renk;
    var rnkLbl=i===0&&!asc?"#1 COK SATAN":"#"+(i+1);
    var rnkGld=isTop?"gld":"";
    var bp=Math.round(Math.max(0,p[st.s]||0)/maxV*100);
    var wid=pfx+"_iw_"+k+"_"+p.code.replace(/[^a-zA-Z0-9]/g,"_");
    var hiBg="background:"+renk+"22;border-color:"+renk;
    var hiSt=function(a){{return a?'style="'+hiBg+'"':"";}};
    var hiSv=function(a){{return a?'style="color:'+renk+';font-size:18px"':"";}};
    var s1=st.s==="top_ytd",s2=st.s==="mag_ytd",s3=st.s==="top_sh"||st.s==="mag_sh";
    var s4=st.s==="onl_ytd"||st.s==="onl_sh",s5=st.s==="str_h",s6=st.s==="str_d";
    var s7=st.s==="ros",s8=st.s==="stok",s9=st.s==="mag_mtd",s10=st.s==="onl_mtd";
    var storyHtml=p.story&&p.story!=="0"&&p.story!==""?'<div class="cstory">'+p.story+"</div>":"";
    card.dataset.sv=String(p[st.s]||0);
    card.innerHTML=
      '<div class="rnk '+rnkGld+'">'+rnkLbl+"</div>"+
      '<div class="ctag">'+p.ust_kat+"</div>"+
      '<div class="iw" id="'+wid+'"><div class="spin"></div></div>'+
      '<div class="cbody">'+
        '<div class="crow"><span class="cgender">'+(gL[p.gender]||p.gender)+"</span>"+
          '<div class="cright"><div class="ckat">'+(p.alt_aile&&p.alt_aile!=="0"?p.alt_aile:p.alt_kat)+"</div>"+
          '<div class="calt">'+p.ana_kat+"</div></div></div>"+
        '<div class="cname">'+(p.aile&&p.aile!=="0"?p.aile:p.code.split(".")[0])+"</div>"+
        storyHtml+
        '<div class="ccode-psf">'+
          '<span class="ccode">'+p.code+"</span>"+
          (p.psf>0?'<span class="cpsf">'+p.psf.toLocaleString("tr")+" TL</span>":"") +
        "</div>"+
        '<div class="sg4">'+
          '<div class="sc" '+hiSt(s1)+'><span class="sl">Top YTD</span><span class="sv" '+hiSv(s1)+">"+p.top_ytd+"</span></div>"+
          '<div class="sc" '+hiSt(s2)+'><span class="sl">Mag YTD</span><span class="sv" '+hiSv(s2)+">"+p.mag_ytd+"</span></div>"+
          '<div class="sc" '+hiSt(s3)+'><span class="sl">Son Hafta</span><span class="sv" '+hiSv(s3)+">"+p.top_sh+"</span></div>"+
          '<div class="sc" '+hiSt(s4)+'><span class="sl">Online YTD</span><span class="sv" '+hiSv(s4)+">"+p.onl_ytd+"</span></div>"+
        "</div>"+
        '<div class="stok-row sg2">'+
          '<div class="sc sc-stok" '+hiSt(s9)+'><span class="sl">MTD Mag</span><span class="sv" '+hiSv(s9)+">"+p.mag_mtd+"</span></div>"+
          '<div class="sc sc-stok" '+hiSt(s10)+'><span class="sl">MTD Online</span><span class="sv" '+hiSv(s10)+">"+p.onl_mtd+"</span></div>"+
        "</div>"+
        ((p.wms>0||p.depo>0)?
          '<div class="wms-row">'+
            '<div class="sc"><span class="sl">WMS Stok</span><span class="sv wms-v">'+p.wms.toLocaleString("tr")+"</span></div>"+
            '<div class="sc"><span class="sl">Depo Stok</span><span class="sv depo-v">'+p.depo.toLocaleString("tr")+"</span></div>"+
            '<div class="sc sc-stok" '+hiSt(s8)+'><span class="sl">Mag Stok</span><span class="sv" '+hiSv(s8)+">"+p.stok.toLocaleString("tr")+"</span></div>"+
          "</div>"
        :
          '<div class="wms-row">'+
            '<div class="sc" style="visibility:hidden"></div>'+
            '<div class="sc" style="visibility:hidden"></div>'+
            '<div class="sc sc-stok" '+hiSt(s8)+'><span class="sl">Mag Stok</span><span class="sv" '+hiSv(s8)+">"+p.stok.toLocaleString("tr")+"</span></div>"+
          "</div>"
        )
        '<div class="str-row">'+
          '<div class="sc3" '+hiSt(s5)+'><span class="sl">STR D.H</span><span class="sv '+sc(p.str_h)+'" '+hiSv(s5)+">"+p.str_h+"%</span></div>"+
          '<div class="sc3" '+hiSt(s6)+'><span class="sl">STR D.D</span><span class="sv '+sc(p.str_d)+'" '+hiSv(s6)+">"+p.str_d+"%</span></div>"+
          '<div class="sc3" '+hiSt(s7)+'><span class="sl">ROS</span><span class="sv" '+hiSv(s7)+">"+p.ros+"</span></div>"+
        "</div>"+
        '<div class="bw"><div class="bf" style="width:'+bp+"%;background:"+renk+'"></div></div>'+
      "</div>";
    grid.appendChild(card);
    (function(src,code,elId){{
      if(!src){{noImg(elId,code);return;}}
      var im=new Image();
      im.onload=function(){{var w=document.getElementById(elId);if(w){{
        im.style.cssText="width:100%;height:100%;object-fit:contain;padding:8px;cursor:zoom-in";
        im.onclick=function(){{window.open(src,"_blank");}};
        w.innerHTML="";w.appendChild(im);
      }}}};
      im.onerror=function(){{noImg(elId,code);}};
      im.src=src;
    }})(p.img,p.code,wid);
  }});
  // Alt toplam
  var totals={{top_ytd:0,mag_ytd:0,onl_ytd:0,top_sh:0,stok:0,mag_mtd:0,onl_mtd:0,wms:0,depo:0}};
  vis.forEach(function(p){{Object.keys(totals).forEach(function(f){{totals[f]+=(p[f]||0);}});}});
  var totRow=document.getElementById(pfx+"-tot-"+k);
  if(totRow){{
    totRow.innerHTML=
      '<div class="tot-cell tot-lbl"><span>TOPLAM</span><span class="tot-n">'+vis.length+' urun</span></div>'+
      '<div class="tot-cell"><span class="tot-l">Top YTD</span><span class="tot-v">'+totals.top_ytd.toLocaleString("tr")+'</span></div>'+
      '<div class="tot-cell"><span class="tot-l">Mag YTD</span><span class="tot-v">'+totals.mag_ytd.toLocaleString("tr")+'</span></div>'+
      '<div class="tot-cell"><span class="tot-l">Online YTD</span><span class="tot-v">'+totals.onl_ytd.toLocaleString("tr")+'</span></div>'+
      '<div class="tot-cell"><span class="tot-l">Son Hafta</span><span class="tot-v">'+totals.top_sh.toLocaleString("tr")+'</span></div>'+
      '<div class="tot-cell"><span class="tot-l">MTD Mag</span><span class="tot-v">'+totals.mag_mtd.toLocaleString("tr")+'</span></div>'+
      '<div class="tot-cell"><span class="tot-l">MTD Onl</span><span class="tot-v">'+totals.onl_mtd.toLocaleString("tr")+'</span></div>'+
      '<div class="tot-cell"><span class="tot-l">Mag Stok</span><span class="tot-v">'+totals.stok.toLocaleString("tr")+'</span></div>'+
      (totals.wms>0?'<div class="tot-cell"><span class="tot-l">WMS</span><span class="tot-v wms-v">'+totals.wms.toLocaleString("tr")+'</span></div>':'')+
      (totals.depo>0?'<div class="tot-cell"><span class="tot-l">Depo</span><span class="tot-v depo-v">'+totals.depo.toLocaleString("tr")+'</span></div>':'');
  }}
  // Sifir filtresi - sort alanina gore gizle
  applyZeroFilter(pfx,k,st.s);
  var lblMap={{top_ytd:"Tum Satislar",mag_ytd:"Magaza YTD",onl_ytd:"Online YTD",
    top_sh:"Son Hafta",mag_sh:"Son Hafta Mag",onl_sh:"Son Hafta Online",
    mag_mtd:"MTD Magaza",onl_mtd:"MTD Online",stok:"Magaza Stok",str_h:"STR",ros:"ROS"}};
  var lbl=document.getElementById(pfx+"-lbl-"+k);
  var cnt=document.getElementById(pfx+"-cnt-"+k);
  if(lbl)lbl.textContent=(asc?"En Az":"En Cok")+" — "+(lblMap[st.s]||st.s);
  if(cnt)cnt.textContent=vis.length+" urun";
}}

var ZERO_STATE={{}};
function applyZeroFilter(pfx,k,field){{
  var fk=pfx+":"+k;
  var showZero=ZERO_STATE[fk]===true;
  var grid=document.getElementById(pfx+"-grid-"+k);
  if(!grid)return;
  var hidden=0,total=0;
  grid.querySelectorAll(".card").forEach(function(card){{
    total++;
    var val=parseFloat(card.dataset.sv)||0;
    if(!showZero&&val===0){{card.style.display="none";hidden++;}}
    else card.style.display="";
  }});
  var btn=document.getElementById(pfx+"-zerobtn-"+k);
  if(!btn)return;
  if(hidden===0&&!showZero){{btn.style.display="none";return;}}
  btn.style.display="flex";
  btn.innerHTML=(showZero?
    '<span class="zarr">&#9660;</span> Hepsini Goster <span class="zbadge">'+total+' urun</span>':
    '<span class="zarr">&#9658;</span> Sifir Satislar <span class="zbadge">'+hidden+' gizli</span>');
  btn.classList.toggle("zero-open",showZero);
}}
function toggleZero(pfx,k){{
  var fk=pfx+":"+k;
  ZERO_STATE[fk]=!ZERO_STATE[fk];
  var st=STATE[fk]||{{}};
  applyZeroFilter(pfx,k,st.s);
}}

function updateKPI(pfx,k){{
  var a=(getAnaliz(pfx,k).toplam)||{{}};
  var KMETA_CUR=window[pfx+"_KMETA"]||KMETA;
  var renk=(KMETA_CUR[k]||{{renk:"#2c2416"}}).renk;
  ["kv1","kv2","kv3","kv4","kv5"].forEach(function(id,i){{
    var el=document.getElementById(pfx+"-"+id+"-"+k);
    if(el)el.textContent=([a.sip_h,a.sip_d,a.mag_ytd,a.onl_ytd,a.top_ytd][i]||0).toLocaleString("en");
  }});
  var kv6=document.getElementById(pfx+"-kv6-"+k);
  var kv7=document.getElementById(pfx+"-kv7-"+k);
  var kv8=document.getElementById(pfx+"-kv8-"+k);
  if(kv6){{kv6.textContent=(a.str_h||0)+"%";kv6.style.color=renk;}}
  if(kv7){{kv7.textContent=(a.str_d||0)+"%";kv7.style.color=renk;}}
  if(kv8)kv8.textContent=(a.stok||0).toLocaleString("en");
}}

function renderStr(pfx,k){{
  var a=getAnaliz(pfx,k);
  var KMETA_CUR=window[pfx+"_KMETA"]||KMETA;
  var renk=(KMETA_CUR[k]||{{renk:"#2c2416"}}).renk;
  var t=a.toplam||{{}};

  function strRow8(d,labelKey,labelVal){{
    var b=Math.min((d.str_h||0)*2,100);
    return "<tr><td style='font-weight:700'>"+labelVal+"</td>"+
      "<td>"+(d.sip_h||0).toLocaleString("tr")+"</td><td>"+(d.sip_d||0).toLocaleString("tr")+"</td>"+
      "<td>"+(d.mag_ytd||0).toLocaleString("tr")+"</td><td>"+(d.onl_ytd||0).toLocaleString("tr")+"</td><td>"+(d.top_ytd||0).toLocaleString("tr")+"</td>"+
      "<td><b class='"+sc(d.str_h||0)+"'>"+(d.str_h||0)+"%</b><div class='sbi' style='width:"+b+"px;background:"+renk+"'></div></td>"+
      "<td><b class='"+sc(d.str_d||0)+"'>"+(d.str_d||0)+"%</b></td></tr>";
  }}

  function totRow8(t,label){{
    var bh=Math.min((t.str_h||0)*2,100);
    return "<tr class='str-tot'>"+
      "<td>"+label+"</td>"+
      "<td>"+(t.sip_h||0).toLocaleString("tr")+"</td><td>"+(t.sip_d||0).toLocaleString("tr")+"</td>"+
      "<td>"+(t.mag_ytd||0).toLocaleString("tr")+"</td><td>"+(t.onl_ytd||0).toLocaleString("tr")+"</td><td>"+(t.top_ytd||0).toLocaleString("tr")+"</td>"+
      "<td><b class='"+sc(t.str_h||0)+"'>"+(t.str_h||0)+"%</b><div class='sbi' style='width:"+bh+"px;background:"+renk+"'></div></td>"+
      "<td><b class='"+sc(t.str_d||0)+"'>"+(t.str_d||0)+"%</b></td></tr>";
  }}

  function aileRow(d){{
    var b=Math.min((d.str_h||0)*2.5,100);
    return "<tr><td style='font-weight:700'>"+d.aile+"</td>"+
      "<td>"+(d.sip_h||0).toLocaleString("tr")+"</td>"+
      "<td>"+(d.mag_ytd||0).toLocaleString("tr")+"</td><td>"+(d.onl_ytd||0).toLocaleString("tr")+"</td><td>"+(d.top_ytd||0).toLocaleString("tr")+"</td><td>"+(d.stok||0).toLocaleString("tr")+"</td>"+
      "<td><b class='"+sc(d.str_h||0)+"'>"+(d.str_h||0)+"%</b><div class='sbi' style='width:"+b+"px;background:"+renk+"'></div></td>"+
      "<td><b class='"+sc(d.str_d||0)+"'>"+(d.str_d||0)+"%</b></td></tr>";
  }}

  // Ana Kategori tablosu
  var kat=a.kat||[];
  var katRows=kat.map(function(d){{return strRow8(d,"kat",d.kat);}}).join("");
  katRows+=totRow8(t,"<b>TOPLAM</b>");
  var el=document.getElementById(pfx+"-st-"+k);
  if(el)el.innerHTML="<div class='stw'><table><thead><tr>"+
    "<th>Kategori</th><th>Sip D.H</th><th>Sip D.D</th><th>Mag YTD</th><th>Online YTD</th><th>Toplam</th><th>STR D.H</th><th>STR D.D</th>"+
    "</tr></thead><tbody>"+katRows+"</tbody></table></div>";

  // Cinsiyet tablosu
  var cin=a.cin||[];
  var cinRows=cin.map(function(d){{return strRow8(d,"cin",gL[d.cin]||d.cin);}}).join("");
  cinRows+=totRow8(t,"<b>TOPLAM</b>");
  var cinEl=document.getElementById(pfx+"-st-cin-"+k);
  if(cinEl)cinEl.innerHTML="<div class='stw'><table><thead><tr>"+
    "<th>Cinsiyet</th><th>Sip D.H</th><th>Sip D.D</th><th>Mag YTD</th><th>Online YTD</th><th>Toplam</th><th>STR D.H</th><th>STR D.D</th>"+
    "</tr></thead><tbody>"+cinRows+"</tbody></table></div>";

  // Ayakkabi Aile tablosu - kendi toplamı
  var aile=a.aile||[];
  var aileTot={{sip_h:0,sip_d:0,mag_ytd:0,onl_ytd:0,top_ytd:0,stok:0,str_h:0,str_d:0}};
  var aileRows=aile.map(function(d){{
    aileTot.sip_h+=(d.sip_h||0);aileTot.sip_d+=(d.sip_d||0);
    aileTot.mag_ytd+=(d.mag_ytd||0);aileTot.onl_ytd+=(d.onl_ytd||0);
    aileTot.top_ytd+=(d.top_ytd||0);aileTot.stok+=(d.stok||0);
    return aileRow(d);
  }}).join("");
  // Aile STR toplam siparis/satis bazlı hesapla
  aileTot.str_h=aileTot.sip_h>0?Math.round(aileTot.top_ytd/aileTot.sip_h*1000)/10:0;
  aileTot.str_d=aileTot.sip_d>0?Math.round(aileTot.top_ytd/aileTot.sip_d*1000)/10:0;
  var aileTotBh=Math.min(aileTot.str_h*2.5,100);
  aileRows+="<tr class='str-tot'><td><b>TOPLAM</b></td>"+
    "<td>"+(aileTot.sip_h).toLocaleString("tr")+"</td>"+
    "<td>"+(aileTot.mag_ytd).toLocaleString("tr")+"</td><td>"+(aileTot.onl_ytd).toLocaleString("tr")+"</td><td>"+(aileTot.top_ytd).toLocaleString("tr")+"</td><td>"+(aileTot.stok).toLocaleString("tr")+"</td>"+
    "<td><b class='"+sc(aileTot.str_h)+"'>"+aileTot.str_h+"%</b><div class='sbi' style='width:"+aileTotBh+"px;background:"+renk+"'></div></td>"+
    "<td><b class='"+sc(aileTot.str_d)+"'>"+aileTot.str_d+"%</b></td></tr>";
  var aileEl=document.getElementById(pfx+"-st-aile-"+k);
  if(aileEl)aileEl.innerHTML="<div class='stw'><table><thead><tr>"+
    "<th>Aile</th><th>Siparis</th><th>Mag YTD</th><th>Online YTD</th><th>Toplam</th><th>Stok</th><th>STR D.H</th><th>STR D.D</th>"+
    "</tr></thead><tbody>"+aileRows+"</tbody></table></div>";
}}

function noImg(wid,code){{
  var w=document.getElementById(wid);if(!w)return;
  w.innerHTML="<div class='noimg'><svg width='24' height='24' viewBox='0 0 24 24' fill='none'><rect x='1' y='1' width='22' height='22' rx='3' stroke='#d4c9b0' stroke-width='1.5'/><path d='M3 18l4-5 3 4 4-5 4 6H3z' stroke='#d4c9b0' stroke-width='1.5' fill='none'/></svg>"+code.split(".")[0]+"</div>";
}}

function switchTab(k,btn,pfx){{
  if(!pfx)pfx=CUR_PFX;
  var tabBar=document.getElementById(pfx+"-tab-bar");
  if(tabBar)tabBar.querySelectorAll(".tab").forEach(function(t){{t.classList.remove("active");}});
  document.querySelectorAll("#sz-wrap-"+pfx+" .panel").forEach(function(p){{p.classList.remove("active");}});
  btn.classList.add("active");
  var panel=document.getElementById(pfx+"-panel-"+k);
  if(panel)panel.classList.add("active");
  initK(pfx,k);
}}

function switchSub(fullKey,s,btn){{
  var parts=fullKey.split("-"),pfx=parts[0],k=parts.slice(1).join("-");
  btn.closest(".sub-tabs").querySelectorAll(".stab").forEach(function(x){{x.classList.remove("active");}});
  btn.classList.add("active");
  var panel=document.getElementById("sp-"+pfx+"-"+k);
  var spPanel=document.getElementById(pfx+"-panel-"+k);
  if(spPanel)spPanel.querySelectorAll(".sub-panel").forEach(function(p){{p.classList.remove("active");}});
  document.getElementById("sp-"+pfx+"-"+k+"-"+s).classList.add("active");
  if(k==="HOSS"){{
    if(s==="s")renderHOSSStr(pfx);
    else if(s==="o")renderHOSSOzet(pfx);
    else renderHOSS(pfx);
  }} else if(s==="s") renderStr(pfx,k);
}}

function switchStrView(fullKey,view,btn){{
  var parts=fullKey.split(":");var pfx=parts[0];var k=parts[1];
  btn.closest(".str-nav").querySelectorAll(".snav").forEach(function(b){{b.classList.remove("act");}});
  btn.classList.add("act");
  var wrap=btn.closest(".str-wrap");
  wrap.querySelectorAll("[id^='ss-']").forEach(function(el){{el.classList.remove("active");}});
  var target=document.getElementById("ss-"+pfx+"-"+k+"-"+view);
  if(target)target.classList.add("active");
  renderStr(pfx,k);
}}

function renderHOSS(pfx){{
  if(!pfx)pfx=CUR_PFX;
  var hs=window[pfx+"_HOSS_STATE"]||(window[pfx+"_HOSS_STATE"]={{f:"all",sort_k:"SSTEP"}});
  var grid=document.getElementById(pfx+"-grid-HOSS");if(!grid)return;
  var data=window[pfx+"_HOSS_URUNLER"]||[];
  var vis=hs.f==="all"?[...data]:data.filter(function(p){{return p.gender===hs.f;}});
  var KMETA_CUR=window[pfx+"_KMETA"]||KMETA;
  grid.innerHTML="";
  vis.forEach(function(p,i){{
    var card=document.createElement("div");
    card.className="card";card.style.borderTop="3px solid #8B5E3C";
    var sk=hs.sort_k;
    var satis=p[sk+"_top_ytd"]||0;var stok=p[sk+"_stok"]||0;var str=p[sk+"_str_h"]||0;var sip=p[sk+"_sip_h"]||0;
    var wid=pfx+"_iw_HOSS_"+p.code.replace(/[^a-zA-Z0-9]/g,"_");
    card.innerHTML='<div class="rnk">#'+(i+1)+"</div>"+
      '<div class="ctag">'+p.ust_kat+"</div>"+
      '<div class="iw" id="'+wid+'"><div class="spin"></div></div>'+
      '<div class="cbody"><div class="crow"><span class="cgender">'+(gL[p.gender]||p.gender)+"</span>"+
        '<div class="cright"><div class="ckat">'+p.ana_kat+"</div></div></div>"+
      '<div class="cname">'+p.ust_kat+"</div>"+
      '<div class="cstory">'+p.story+"</div>"+
      '<div class="ccode">'+p.code+"</div>"+
      '<div class="sg4">'+
        '<div class="sc"><span class="sl">Siparis</span><span class="sv">'+sip+"</span></div>"+
        '<div class="sc hi"><span class="sl">Satis ('+sk+')</span><span class="sv">'+satis+"</span></div>"+
        '<div class="sc"><span class="sl">Stok</span><span class="sv">'+stok+"</span></div>"+
        '<div class="sc"><span class="sl">STR</span><span class="sv '+sc(str)+'">'+str+"%</span></div>"+
      "</div></div>";
    grid.appendChild(card);
    (function(src,code,elId){{
      if(!src){{noImg(elId,code);return;}}
      var im=new Image();
      im.onload=function(){{var w=document.getElementById(elId);if(w){{im.style.cssText="width:100%;height:100%;object-fit:contain;padding:8px;cursor:zoom-in";im.onclick=function(){{window.open(src,"_blank");}};w.innerHTML="";w.appendChild(im);}}}};
      im.onerror=function(){{noImg(elId,code);}};im.src=src;
    }})(p.img,p.code,wid);
  }});
  var lbl=document.getElementById(pfx+"-lbl-HOSS");
  if(lbl)lbl.textContent=hs.sort_k+" Satisina Gore";
  var cnt=document.getElementById(pfx+"-cnt-HOSS");
  if(cnt)cnt.textContent=vis.length+" urun";
}}

function renderHOSSStr(pfx){{
  if(!pfx)pfx=CUR_PFX;
  var KANALLER=Object.keys(window[pfx+"_KMETA"]||KMETA);
  function hossTable(data,labelKey){{
    var rows="";
    data.forEach(function(d){{
      rows+="<tr><td style='font-weight:700'>"+d[labelKey]+"</td>";
      KANALLER.forEach(function(k){{
        rows+="<td>"+(d[k+"_sip_h"]||0)+"</td><td>"+(d[k+"_top_ytd"]||0)+"</td><td><b class='"+sc(d[k+"_str_h"]||0)+"'>"+(d[k+"_str_h"]||0)+"%</b></td>";
      }});
      rows+="</tr>";
    }});
    var hdr="<tr><th>"+labelKey+"</th>";
    KANALLER.forEach(function(k){{hdr+="<th>"+k+" Sip</th><th>"+k+" Satis</th><th>"+k+" STR</th>";}});
    hdr+="</tr>";
    return "<div class='stw'><table><thead>"+hdr+"</thead><tbody>"+rows+"</tbody></table></div>";
  }}
  var katEl=document.getElementById(pfx+"-st-HOSS-kat");
  if(katEl)katEl.innerHTML=hossTable(window[pfx+"_HOSS_KAT"]||[],"kat");
  var ustEl=document.getElementById(pfx+"-st-HOSS-ust");
  if(ustEl)ustEl.innerHTML=hossTable(window[pfx+"_HOSS_UST"]||[],"ust");
  var kodEl=document.getElementById(pfx+"-st-HOSS-kod");
  if(kodEl){{
    var rows="";
    (window[pfx+"_HOSS_URUNLER"]||[]).forEach(function(p){{
      rows+="<tr><td style='font-family:monospace;font-weight:600'>"+p.code+"</td><td>"+p.gender+"</td><td>"+p.ana_kat+"</td><td>"+p.ust_kat+"</td>";
      KANALLER.forEach(function(k){{rows+="<td>"+(p[k+"_sip_h"]||0)+"</td><td>"+(p[k+"_top_ytd"]||0)+"</td><td>"+(p[k+"_stok"]||0)+"</td><td><b class='"+sc(p[k+"_str_h"]||0)+"'>"+(p[k+"_str_h"]||0)+"%</b></td>";}});
      rows+="</tr>";
    }});
    var hdr="<tr><th>Kod</th><th>Cin</th><th>Ana Kat</th><th>Ust Kat</th>";
    KANALLER.forEach(function(k){{hdr+="<th>"+k+" Sip</th><th>"+k+" Satis</th><th>"+k+" Stok</th><th>"+k+" STR</th>";}});
    kodEl.innerHTML="<div class='stw'><table><thead>"+hdr+"</thead><tbody>"+rows+"</tbody></table></div>";
  }}
}}

function renderHOSSOzet(pfx){{
  if(!pfx)pfx=CUR_PFX;
  var KANALLER=Object.keys(window[pfx+"_KMETA"]||KMETA);
  var cinEl=document.getElementById(pfx+"-st-HOSS-cin");if(!cinEl)return;
  var rows="";
  (window[pfx+"_HOSS_CIN"]||[]).forEach(function(d){{
    rows+="<tr><td style='font-weight:700'>"+(gL[d.cin]||d.cin)+"</td>";
    KANALLER.forEach(function(k){{rows+="<td>"+(d[k+"_sip_h"]||0)+"</td><td>"+(d[k+"_top_ytd"]||0)+"</td><td>"+(d[k+"_stok"]||0)+"</td><td><b class='"+sc(d[k+"_str_h"]||0)+"'>"+(d[k+"_str_h"]||0)+"%</b></td>";}});
    rows+="</tr>";
  }});
  var hdr="<tr><th>Cinsiyet</th>";
  KANALLER.forEach(function(k){{hdr+="<th>"+k+" Sip</th><th>"+k+" Satis</th><th>"+k+" Stok</th><th>"+k+" STR</th>";}});
  cinEl.innerHTML="<div class='stw'><table><thead>"+hdr+"</thead><tbody>"+rows+"</tbody></table></div>";
}}

var HOSS_INITED={{}};
var inited={{}};
function initK(pfx,k){{
  var fk=pfx+":"+k;
  if(inited[fk])return;inited[fk]=true;
  updateKPI(pfx,k);
  setupTB(pfx,k);
  renderGrid(fk);
  if(k==="HOSS"){{
    if(!HOSS_INITED[pfx]){{
      HOSS_INITED[pfx]=true;
      window[pfx+"_HOSS_STATE"]={{f:"all",sort_k:"SSTEP"}};
      var sg=document.getElementById(pfx+"-sg-HOSS");
      if(sg)sg.addEventListener("click",function(e){{
        var b=e.target.closest("[data-k]");if(!b)return;
        sg.querySelectorAll("[data-k]").forEach(function(x){{x.classList.remove("as");}});
        b.classList.add("as");window[pfx+"_HOSS_STATE"].sort_k=b.dataset.k;renderHOSS(pfx);
      }});
      var fg=document.getElementById(pfx+"-fg-HOSS");
      if(fg)fg.addEventListener("click",function(e){{
        var b=e.target.closest("[data-f]");if(!b)return;
        fg.querySelectorAll("[data-f]").forEach(function(x){{x.classList.remove("af");}});
        b.classList.add("af");window[pfx+"_HOSS_STATE"].f=b.dataset.f;renderHOSS(pfx);
      }});
    }}
    renderHOSS(pfx);
  }}
}}

function setupTB(pfx,k){{
  var fk=pfx+":"+k;
  var sgEl=document.getElementById(pfx+"-sg-"+k);
  if(sgEl)sgEl.addEventListener("click",function(e){{
    var b=e.target.closest("[data-s]");if(!b)return;
    sgEl.querySelectorAll("[data-s]").forEach(function(x){{x.classList.remove("as");}});
    b.classList.add("as");
    if(!STATE[fk])STATE[fk]={{f:"all",kat:"all",aile:[],s:"top_ytd",d:"desc"}};
    STATE[fk].s=b.dataset.s;STATE[fk].d=b.dataset.d||"desc";
    renderGrid(fk);
  }});
  var aileListEl=document.getElementById(pfx+"-aile-list-"+k);
  if(aileListEl)aileListEl.addEventListener("change",function(){{aileChange(pfx+":"+k);}});
}}

function aileChange(fullKey){{
  var parts=fullKey.split(":");var pfx=parts[0];var k=parts[1];
  var checks=document.querySelectorAll("#"+pfx+"-aile-list-"+k+" input:checked");
  if(!STATE[fullKey])STATE[fullKey]={{f:"all",kat:"all",aile:[],s:"top_ytd",d:"desc"}};
  STATE[fullKey].aile=[].slice.call(checks).map(function(c){{return c.value;}});
  var val=document.getElementById(pfx+"-ddv-aile-"+k);
  if(val)val.textContent=STATE[fullKey].aile.length?(STATE[fullKey].aile.length+" secili"):"Tumu";
  var btn=document.getElementById(pfx+"-ddb-aile-"+k);
  if(btn){{STATE[fullKey].aile.length?btn.classList.add("active"):btn.classList.remove("active");}}
  renderGrid(fullKey);
}}
function aileTumunu(fullKey){{
  var parts=fullKey.split(":");var pfx=parts[0];var k=parts[1];
  document.querySelectorAll("#"+pfx+"-aile-list-"+k+" input").forEach(function(c){{c.checked=true;}});
  aileChange(fullKey);
}}
function aileTemizle(fullKey){{
  var parts=fullKey.split(":");var pfx=parts[0];var k=parts[1];
  document.querySelectorAll("#"+pfx+"-aile-list-"+k+" input").forEach(function(c){{c.checked=false;}});
  if(STATE[fullKey])STATE[fullKey].aile=[];
  var val=document.getElementById(pfx+"-ddv-aile-"+k);if(val)val.textContent="Tumu";
  var btn=document.getElementById(pfx+"-ddb-aile-"+k);if(btn)btn.classList.remove("active");
  renderGrid(fullKey);
}}

function switchSezon(s){{
  document.querySelectorAll(".sezon-btn").forEach(function(b){{b.classList.remove("active");}});
  document.getElementById("sz-btn-"+s).classList.add("active");
  document.querySelectorAll(".sezon-wrap").forEach(function(w){{w.classList.remove("active");}});
  document.getElementById("sz-wrap-"+s).classList.add("active");
  CUR_PFX=s;
  var firstTab=document.querySelector("#sz-wrap-"+s+" .tab");
  if(firstTab)firstTab.click();
}}

function indir(k){{
  if(typeof sendPrompt==="function")sendPrompt(k+" gitmeyen Excel indir");
  else alert("python lacoste_rapor_v2.py --indir");
}}

document.addEventListener("click",function(e){{
  if(!e.target.closest(".dd-wrap")){{
    document.querySelectorAll(".dd-menu.open").forEach(function(m){{m.classList.remove("open");}});
    document.querySelectorAll(".dd-btn.active").forEach(function(b){{b.classList.remove("active");}});
  }}
}});

window["AW_KMETA"]=AW_KMETA;
window["SS_KMETA"]=SS_KMETA;
setTimeout(function(){{
  var t=document.querySelector("#sz-wrap-AW .tab");
  if(t){{t.click();}}
}},50);
// Fallback: sayfa zaten yüklendiyse
if(document.readyState==="complete"||document.readyState==="interactive"){{
  setTimeout(function(){{
    var t=document.querySelector("#sz-wrap-AW .tab");
    if(t)t.click();
  }},50);
}}
</script>
</body>
</html>"""
    return _html

def uret_html(df, hafta_str, excel_adi):
    J = lambda o: json.dumps(o, ensure_ascii=True)
    total = len(df)

    kanal_data = {}
    for kanal in KANALLAR:
        kanal_data[kanal] = {
            "urunler": urun_listesi(df, kanal),
            "analiz":  str_analiz(df, kanal),
            "git_cnt": len(gitmeyen_df(df, kanal)),
        }
    hoss_data = hoss_analiz(df)

    def urun_js(rows):
        items = [
            (f'{{code:{J(p["code"])},gender:{J(p["gender"])},ana_kat:{J(p["ana_kat"])},'
             f'ust_kat:{J(p["ust_kat"])},alt_kat:{J(p["alt_kat"])},aile:{J(p["aile"])},'
             f'alt_aile:{J(p["alt_aile"])},sip_grp:{J(p["sip_grp"])},'
             f'sip_h:{p["sip_h"]},sip_d:{p["sip_d"]},'
             f'mag_ytd:{p["mag_ytd"]},onl_ytd:{p["onl_ytd"]},top_ytd:{p["top_ytd"]},'
             f'mag_sh:{p["mag_sh"]},onl_sh:{p["onl_sh"]},top_sh:{p["top_sh"]},'
             f'stok:{p["stok"]},ros:{p["ros"]},str_h:{p["str_h"]},str_d:{p["str_d"]},'
             f'story:{J(p.get("story",""))},is_hoss:{"true" if p.get("is_hoss") else "false"},'
             f'mag_mtd:{p.get("mag_mtd",0)},onl_mtd:{p.get("onl_mtd",0)},top_mtd:{p.get("top_mtd",0)},'
             f'psf:{p.get("psf",0)},'
             f'img:{J(p["img"])}}}')
            for p in rows
        ]
        return "[\n" + ",\n".join(items) + "\n]"

    # const yerine var kullan - tarayicida window["X"] ile erisim icin
    js_data = "\n".join(
        f'var URUNLER_{k}={urun_js(kanal_data[k]["urunler"])};'
        for k in KANALLAR
    )
    js_analiz = "\n".join(
        f'var ANALIZ_{k}={J(kanal_data[k]["analiz"])};'
        for k in KANALLAR
    )
    kmeta = J({k: {"renk": KANALLAR[k]["renk"], "git_cnt": kanal_data[k]["git_cnt"]} for k in KANALLAR})

    # HOSS JS degiskenleri
    hoss_js_urunler = _hoss_urun_js(J, hoss_data.get("urunler", []))
    hoss_js_kat  = J(hoss_data.get("ozet_kat", []))
    hoss_js_ust  = J(hoss_data.get("ozet_ust", []))
    hoss_js_cin  = J(hoss_data.get("ozet_cin", []))
    hoss_count   = hoss_data.get("count", 0)

    tabs = "\n".join(
        f'<button class="tab" data-renk="{KANALLAR[k]["renk"]}" onclick="switchTab(\'{k}\',this)" style="--tk:{KANALLAR[k]["renk"]}">{k} <span class="tc">{kanal_data[k]["git_cnt"]} bekleyen</span></button>'
        for k in KANALLAR
    )
    tabs += f'\n<button class="tab hoss-tab" data-renk="#7b3f00" style="--tk:#7b3f00" onclick="switchTab(\'HOSS\',this)">&#x2726; HOSS x Lacoste <span class="tc">{hoss_count} urun</span></button>'

    panels = ""
    for k, cfg in KANALLAR.items():
        git = kanal_data[k]["git_cnt"]
        panels += f"""
<div class="panel" id="panel-{k}">
  <div class="kpi-bar" style="border-bottom:3px solid {cfg['renk']}">
    <div class="kpi"><div class="kpi-l">Sip. Devir Haric</div><div class="kpi-v" id="kv1-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Sip. Devir Dahil</div><div class="kpi-v" id="kv2-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Mag. Satis YTD</div><div class="kpi-v" id="kv3-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Online Satis YTD</div><div class="kpi-v" id="kv4-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Toplam Satis YTD</div><div class="kpi-v" id="kv5-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">STR Devir Haric</div><div class="kpi-v" style="color:{cfg['renk']};font-size:24px" id="kv6-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">STR Devir Dahil</div><div class="kpi-v" style="color:{cfg['renk']};font-size:24px" id="kv7-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Kanal Stok</div><div class="kpi-v" id="kv8-{k}">-</div></div><div class="kdiv"></div>
    <div class="kpi"><div class="kpi-l">Magazaya Gitmeyen</div>
      <div class="kpi-v">{git} <button class="dlb" onclick="indir('{k}')">Excel</button></div>
    </div>
  </div>
  <div class="sub-tabs">
    <button class="stab active" onclick="switchSub('{k}','u',this)">Tum Urunler</button>
    <button class="stab" onclick="switchSub('{k}','s',this)">STR Analizi</button>
  </div>
  <div class="sub-panel active" id="sp-{k}-u">
    <div class="toolbar2">
      <div class="search-bar">
        <div class="srch-wrap srch-wrap-rel" id="SGL-srch-wrap-{k}">
          <span class="srch-icon">&#128269;</span>
          <input class="srch-inp" id="SGL-srch-{k}" type="text" placeholder="Option ara... (Enter = Tümünü Seç)"
            oninput="onSearch('SGL:{k}',this.value)"
            onclick="event.stopPropagation();showSearchDropdown('SGL:{k}',this.value)"
            onfocus="showSearchDropdown('SGL:{k}',this.value)">
          <button class="srch-clear" id="SGL-srch-clear-{k}" onclick="clearSearch('SGL:{k}')" style="display:none">&#10005;</button>
        </div>
        <div class="sel-chips" id="SGL-chips-{k}"></div>
      </div>
      <div class="tb-left">
        <div class="dd-wrap" id="ddw-cin-{k}">
          <button class="dd-btn" id="ddb-cin-{k}" onclick="toggleDD('{k}','cin')">
            <i class="ti ti-user" aria-hidden="true"></i> Cinsiyet <span class="dd-val" id="ddv-cin-{k}">Tumu</span> <i class="ti ti-chevron-down" aria-hidden="true"></i>
          </button>
          <div class="dd-menu" id="ddm-cin-{k}">
            <label class="dd-item"><input type="radio" name="cin-{k}" value="all" checked onchange="ddChange('{k}','cin',this)"> Tumu</label>
            <label class="dd-item"><input type="radio" name="cin-{k}" value="MEN" onchange="ddChange('{k}','cin',this)"> Erkek</label>
            <label class="dd-item"><input type="radio" name="cin-{k}" value="WOMEN" onchange="ddChange('{k}','cin',this)"> Kadin</label>
            <label class="dd-item"><input type="radio" name="cin-{k}" value="CHILDREN" onchange="ddChange('{k}','cin',this)"> Cocuk</label>
            <label class="dd-item"><input type="radio" name="cin-{k}" value="ACCESSORIES" onchange="ddChange('{k}','cin',this)"> Aksesuar</label>
          </div>
        </div>
        <div class="dd-wrap" id="ddw-kat-{k}">
          <button class="dd-btn" id="ddb-kat-{k}" onclick="toggleDD('{k}','kat')">
            <i class="ti ti-tag" aria-hidden="true"></i> Kategori <span class="dd-val" id="ddv-kat-{k}">Tumu</span> <i class="ti ti-chevron-down" aria-hidden="true"></i>
          </button>
          <div class="dd-menu" id="ddm-kat-{k}">
            <label class="dd-item"><input type="radio" name="kat-{k}" value="all" checked onchange="ddChange('{k}','kat',this)"> Tumu</label>
            <label class="dd-item"><input type="radio" name="kat-{k}" value="FOOTWEAR" onchange="ddChange('{k}','kat',this)"> Footwear</label>
            <label class="dd-item"><input type="radio" name="kat-{k}" value="TEXTILE" onchange="ddChange('{k}','kat',this)"> Textile</label>
            <label class="dd-item"><input type="radio" name="kat-{k}" value="TEXTILE ACCESSORIES" onchange="ddChange('{k}','kat',this)"> Textile Acc.</label>
            <label class="dd-item"><input type="radio" name="kat-{k}" value="ACCESSORIES" onchange="ddChange('{k}','kat',this)"> Accessories</label>
            <label class="dd-item"><input type="radio" name="kat-{k}" value="UNDERWEAR" onchange="ddChange('{k}','kat',this)"> Underwear</label>
          </div>
        </div>
        <div class="dd-wrap" id="ddw-aile-{k}">
          <button class="dd-btn" id="ddb-aile-{k}" onclick="toggleDD('{k}','aile')">
            <i class="ti ti-shoe" aria-hidden="true"></i> Aile <span class="dd-val" id="ddv-aile-{k}">Tumu</span> <i class="ti ti-chevron-down" aria-hidden="true"></i>
          </button>
          <div class="dd-menu dd-multi" id="ddm-aile-{k}">
            <div class="dd-multi-hdr">Coklu secim</div>
            <div id="aile-list-{k}"></div>
            <div class="dd-multi-footer">
              <button class="dd-act" onclick="aileTumunu('{k}')">Tumu</button>
              <button class="dd-act" onclick="aileTemizle('{k}')">Temizle</button>
            </div>
          </div>
        </div>
      </div>
      <div class="tb-right">
        <span class="tlbl2">Sirala</span>
        <div id="sg-{k}" class="sort-group">
          <button class="sbtn as" data-s="top_ytd" data-d="desc">Tum Satislar</button>
          <button class="sbtn" data-s="mag_ytd" data-d="desc">Magaza YTD</button>
          <button class="sbtn" data-s="onl_ytd" data-d="desc">Online YTD</button>
          <button class="sbtn" data-s="top_sh" data-d="desc">Son Hafta Mag</button>
          <button class="sbtn" data-s="onl_sh" data-d="desc">Son Hafta Online</button>
          <button class="sbtn" data-s="mag_mtd" data-d="desc">MTD Mag</button>
          <button class="sbtn" data-s="onl_mtd" data-d="desc">MTD Online</button>
          <button class="sbtn" data-s="stok" data-d="desc">Magaza Stok</button>
          <button class="sbtn" data-s="str_h" data-d="desc">STR</button>
          <button class="sbtn" data-s="ros" data-d="desc">ROS</button>
          <button class="sbtn" data-s="top_ytd" data-d="asc">En Az Satanlar</button>
        </div>
      </div>
    </div>
    <div class="gw">
      <div class="shdr"><span id="lbl-{k}">Toplam Satisa Gore</span><span id="cnt-{k}">-</span></div>
      <div class="grid" id="grid-{k}"></div>
    </div>
  </div>
  <div class="sub-panel" id="sp-{k}-s">
    <div class="str-wrap">
      <div style="display:flex;gap:8px;margin-bottom:1rem;flex-wrap:wrap">
        <button class="str-nav-btn active" onclick="switchStrView('{k}','kat',this)">Ana Kategori</button>
        <button class="str-nav-btn" onclick="switchStrView('{k}','cin',this)">Cinsiyet</button>
        <button class="str-nav-btn" onclick="switchStrView('{k}','aile',this)">Urun Ailesi</button>
      </div>
      <div class="str-sub active" id="str-kat-view-{k}">
        <div class="str-section">
          <div class="str-title" style="border-color:{cfg['renk']}">Ana Kategori Bazli STR</div>
          <div class="stw" id="st-{k}"></div>
        </div>
      </div>
      <div class="str-sub" id="str-cin-view-{k}">
        <div class="str-section">
          <div class="str-title" style="border-color:{cfg['renk']}">Cinsiyet Bazli STR</div>
          <div class="stw" id="st-cin-{k}"></div>
        </div>
      </div>
      <div class="str-sub" id="str-aile-view-{k}">
        <div class="str-section">
          <div class="str-title" style="border-color:{cfg['renk']}">Urun Ailesi Bazli STR</div>
          <div class="stw" id="st-aile-{k}"></div>
        </div>
      </div>
    </div>
  </div>
</div>"""


    hoss_panel = f"""
<div class="panel" id="panel-HOSS">
  <div class="hoss-kpi">
    <div class="kpi"><div class="kpi-l">Toplam Urun</div><div class="kpi-v">{hoss_count}</div></div>
    <div class="kdiv" style="background:rgba(196,122,58,.3)"></div>
    <div class="kpi"><div class="kpi-l">Koleksiyon</div><div class="kpi-v" style="font-size:14px">HOSS x Lacoste Kapsul</div></div>
  </div>
  <div class="sub-tabs">
    <button class="stab active" onclick="switchSub('HOSS','urunler',this)">Tum Urunler</button>
    <button class="stab" onclick="switchSub('HOSS','str',this)">STR Analizi</button>
    <button class="stab" onclick="switchSub('HOSS','ozet',this)">Ozet Tablo</button>
  </div>
  <div class="sub-panel active" id="sp-HOSS-urunler">
    <div class="toolbar">
      <span class="tlbl">Filtre</span>
      <div id="fg-HOSS">
        <button class="btn af" data-f="all">Tumu</button>
        <button class="btn" data-f="MEN">Erkek</button>
        <button class="btn" data-f="WOMEN">Kadin</button>
        <button class="btn" data-f="ACCESSORIES">Aksesuar</button>
      </div>
      <div class="tsep"></div>
      <span class="tlbl">Kanal Siralama</span>
      <div id="sg-HOSS">
        {''.join([f'<button class="btn{" as" if i==0 else ""}" data-k="{k}">{k}</button>' for i,k in enumerate(KANALLAR)])}
      </div>
    </div>
    <div class="gw">
      <div class="shdr"><span id="lbl-HOSS">SSTEP Satisina Gore</span><span id="cnt-HOSS">-</span></div>
      <div class="grid" id="grid-HOSS"></div>
    </div>
  </div>
  <div class="sub-panel" id="sp-HOSS-str">
    <div class="str-wrap">
      <div class="str-section">
        <div class="str-title" style="border-color:#7b3f00">Ana Kategori Bazli STR (Tum Kanallar)</div>
        <div class="stw" id="st-HOSS-kat"></div>
      </div>
      <div class="str-section">
        <div class="str-title" style="border-color:#7b3f00">Ust Kategori Bazli STR</div>
        <div class="stw" id="st-HOSS-ust"></div>
      </div>
      <div class="str-section">
        <div class="str-title" style="border-color:#7b3f00">Kod Bazli STR (Tum Kanallar)</div>
        <div class="stw" id="st-HOSS-kod"></div>
      </div>
    </div>
  </div>
  <div class="sub-panel" id="sp-HOSS-ozet">
    <div class="str-wrap">
      <div class="str-section">
        <div class="str-title" style="border-color:#7b3f00">Cinsiyet Bazli Ozet</div>
        <div class="stw" id="st-HOSS-cin"></div>
      </div>
    </div>
  </div>
</div>"""

    return f"""<!DOCTYPE html>
<html lang="tr"><head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Lacoste 26AW Takip — {hafta_str}</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700&family=DM+Sans:wght@400;500;600;700&display=swap');
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --bg:#f5f0e8;--bg2:#ede8de;--bg3:#e4ddd0;--bg4:#d8d0c0;
  --card:#ffffff;--dark:#2c2416;--dark2:#3a2e1a;--dark3:#5a4a2a;
  --gold:#C8A84B;--gold2:#a08030;--gold3:#f0e4c0;
  --border:#d4c9b0;--border2:#c4b99a;
  --text:#2c2416;--text2:#6a5a3a;--text3:#9a8a6a;
  --r:10px;--r2:14px;
}}
body{{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--text);min-height:100vh}}

/* HEADER */
.header{{background:var(--dark);padding:.8rem 1rem;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:.5rem}}
.logo-wrap{{display:flex;align-items:center;gap:10px}}
.logo-icon{{width:34px;height:34px;border-radius:8px;background:var(--gold);display:flex;align-items:center;justify-content:center;font-size:16px;flex-shrink:0}}
.brand{{font-family:'Playfair Display',serif;font-size:1rem;letter-spacing:1.5px;color:#f0e8d0;text-transform:uppercase;line-height:1.2}}
.brand-sub{{font-size:9px;color:#6a5a3a;letter-spacing:1.5px;text-transform:uppercase;margin-top:2px}}
.hb{{font-size:9px;font-weight:600;padding:4px 10px;border-radius:16px;background:#3a2e1a;color:var(--gold);border:1px solid #4a3c22;white-space:nowrap}}

/* TABS */
.tab-bar{{background:var(--bg2);border-bottom:1px solid var(--border);padding:0 1rem;display:flex;overflow-x:auto;-webkit-overflow-scrolling:touch;scrollbar-width:none}}
.tab-bar::-webkit-scrollbar{{display:none}}
.tab{{padding:.55rem .9rem;font-size:11px;font-weight:600;cursor:pointer;border:none;background:none;color:var(--text3);border-bottom:3px solid transparent;margin-bottom:-1px;font-family:'DM Sans',sans-serif;white-space:nowrap;transition:all .2s}}
.tab.active{{color:var(--tk,var(--dark));border-bottom-color:var(--tk,var(--gold));font-weight:700}}
.tab:hover{{color:var(--dark)}}
.hoss-tab{{color:#8B5E3C}}.hoss-tab.active{{border-bottom-color:#8B5E3C;color:#8B5E3C}}
.tc{{font-size:9px;margin-left:3px;opacity:.5;font-weight:400}}

/* KPI BAR */
.kpi-bar{{background:var(--bg2);padding:.5rem 1rem;display:grid;grid-template-columns:repeat(auto-fit,minmax(90px,1fr));gap:.4rem;border-bottom:1px solid var(--border)}}
.kpi{{display:flex;flex-direction:column;gap:1px;padding:.3rem .4rem;background:var(--card);border-radius:6px;border:1px solid var(--border)}}
.kpi-l{{font-size:7.5px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:var(--text3);line-height:1.3}}
.kpi-v{{font-family:'Playfair Display',serif;font-size:17px;color:var(--dark);line-height:1;display:flex;align-items:center;gap:5px;flex-wrap:wrap}}
.kdiv{{display:none}}
.dlb{{font-size:9px;font-weight:700;padding:3px 8px;border-radius:6px;border:1px solid var(--border2);background:var(--bg);color:var(--text2);cursor:pointer;white-space:nowrap}}

/* SUB TABS */
.sub-tabs{{background:var(--bg);border-bottom:1px solid var(--border);padding:0 1rem;display:flex}}
.stab{{padding:.45rem .8rem;font-size:11px;font-weight:600;cursor:pointer;border:none;background:none;color:var(--text3);border-bottom:2px solid transparent;margin-bottom:-1px;font-family:'DM Sans',sans-serif;transition:all .15s}}
.stab.active{{color:var(--dark);border-bottom-color:var(--border2)}}

/* TOOLBAR */
.toolbar2{{background:var(--bg);border-bottom:1px solid var(--border);padding:.4rem 1rem;display:flex;flex-direction:column;gap:.4rem}}
.tb-left{{display:flex;align-items:center;gap:5px;flex-wrap:wrap}}
.tb-right{{display:flex;align-items:center;gap:4px;flex-wrap:wrap}}
.tlbl2{{font-size:9px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.8px;white-space:nowrap}}

/* DROPDOWN */
.dd-wrap{{position:relative}}
.dd-btn{{display:flex;align-items:center;gap:4px;border:1px solid var(--border2);border-radius:8px;padding:5px 9px;font-size:11px;font-weight:600;cursor:pointer;color:var(--text2);background:var(--bg2);font-family:'DM Sans',sans-serif;white-space:nowrap;transition:all .15s;min-height:32px}}
.dd-btn:hover,.dd-btn.active{{background:var(--dark);color:#f0e8d0;border-color:var(--dark)}}
.dd-val{{font-size:10px;font-weight:400;color:var(--text3);margin-left:1px}}
.dd-btn.active .dd-val{{color:#a08050}}
.dd-menu{{display:none;position:fixed;background:var(--card);border:1px solid var(--border2);border-radius:var(--r2);padding:6px;min-width:160px;z-index:1000;box-shadow:0 8px 24px rgba(44,36,22,.18);max-height:50vh;overflow-y:auto}}
.dd-menu.open{{display:block}}
.dd-item{{display:flex;align-items:center;gap:7px;padding:7px 8px;cursor:pointer;font-size:12px;color:var(--text);border-radius:6px;transition:background .12s}}
.dd-item:hover{{background:var(--bg2)}}
.dd-item input{{cursor:pointer;accent-color:var(--tk,var(--dark));width:14px;height:14px;flex-shrink:0}}
.dd-multi{{min-width:200px;max-height:50vh;display:none;flex-direction:column}}
.dd-multi.open{{display:flex}}
.dd-multi-hdr{{font-size:9px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.8px;padding:4px 8px 6px;border-bottom:1px solid var(--border);flex-shrink:0}}
.dd-multi>div:nth-child(2){{overflow-y:auto;flex:1}}
.dd-multi-footer{{display:flex;gap:5px;padding:6px 8px 2px;border-top:1px solid var(--border);flex-shrink:0}}
.dd-act{{font-size:10px;font-weight:600;padding:4px 9px;border-radius:6px;border:1px solid var(--border2);background:var(--bg2);color:var(--text2);cursor:pointer;font-family:'DM Sans',sans-serif}}
.dd-act:hover{{background:var(--dark);color:#f0e8d0}}

/* SORT BUTTONS */
.sort-group{{display:flex;gap:3px;flex-wrap:wrap}}
.sbtn{{border:1px solid var(--border);border-radius:6px;padding:4px 8px;font-size:10px;font-weight:600;cursor:pointer;color:var(--text2);background:var(--bg2);font-family:'DM Sans',sans-serif;white-space:nowrap;transition:all .15s;min-height:28px}}
.sbtn:hover{{background:var(--bg3);color:var(--dark)}}
.sbtn.as{{background:var(--tk,var(--dark));color:#fff;border-color:var(--tk,var(--dark))}}
.btn{{border:1px solid var(--border2);border-radius:20px;padding:3px 10px;font-size:11px;font-weight:600;cursor:pointer;color:var(--text2);background:var(--bg2);font-family:'DM Sans',sans-serif;white-space:nowrap;transition:all .15s}}
.btn.af{{background:var(--dark);color:#f0e8d0;border-color:var(--dark)}}
.btn.as{{background:var(--tk,var(--dark));color:#fff;border-color:var(--tk,var(--dark))}}
.tlbl{{font-size:9px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:1px}}
.tsep{{width:1px;height:16px;background:var(--border);margin:0 3px}}

/* GRID */
.gw{{padding:.8rem 1rem 2rem}}
.shdr{{font-size:9px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:var(--text3);margin-bottom:.8rem;padding-bottom:5px;border-bottom:1px solid var(--border);display:flex;justify-content:space-between;align-items:center}}
.grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(155px,1fr));gap:8px}}

/* CARDS */
.card{{background:var(--card);border-radius:var(--r2);overflow:hidden;border:1px solid var(--border);display:flex;flex-direction:column;transition:border-color .2s,box-shadow .2s;position:relative}}
.card:hover{{border-color:var(--border2);box-shadow:0 3px 12px rgba(44,36,22,.07)}}
.card.top{{border-color:var(--gold);border-width:1.5px}}
.rnk{{position:absolute;top:6px;left:6px;font-size:8px;font-weight:700;padding:2px 6px;border-radius:5px;z-index:4;background:var(--dark);color:#f0e8d0;line-height:1.4}}
.rnk.gld{{background:var(--gold);color:var(--dark)}}
.ctag{{position:absolute;top:6px;right:6px;font-size:7px;font-weight:700;text-transform:uppercase;padding:2px 5px;border-radius:4px;z-index:4;background:var(--bg2);color:var(--text3);max-width:70px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.iw{{width:100%;height:140px;background:var(--bg);display:flex;align-items:center;justify-content:center;overflow:hidden;flex-shrink:0;border-bottom:1px solid var(--bg2)}}
.iw img{{width:100%;height:100%;object-fit:contain;padding:6px;transition:transform .25s}}
.spin{{width:20px;height:20px;border:2px solid var(--border2);border-top-color:var(--gold);border-radius:50%;animation:sp .7s linear infinite}}
@keyframes sp{{to{{transform:rotate(360deg)}}}}
.cbody{{padding:7px 8px 8px;flex:1;display:flex;flex-direction:column;gap:3px}}
.crow{{display:flex;justify-content:space-between;align-items:flex-start;gap:2px}}
.cgender{{font-size:7px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--text3);flex-shrink:0}}
.cright{{display:flex;flex-direction:column;align-items:flex-end}}
.ckat{{font-size:7.5px;font-weight:600;color:var(--text3);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:85px}}
.calt{{font-size:7px;color:var(--text3);opacity:.7;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:85px}}
.cname{{font-size:10.5px;font-weight:700;color:var(--dark);line-height:1.3;min-height:22px}}
.cstory{{font-size:7.5px;color:#7a4e2a;font-weight:600;background:var(--gold3);border-radius:3px;padding:2px 5px;border-left:2px solid var(--gold);font-style:italic;margin-top:-1px}}
.ccode-psf{{display:flex;align-items:center;justify-content:space-between;gap:4px;flex-wrap:wrap}}
.ccode{{font-size:8px;font-family:monospace;background:var(--bg2);color:var(--text2);padding:2px 5px;border-radius:3px;border:1px solid var(--border);align-self:flex-start;font-weight:600}}
.cpsf{{font-size:9px;font-weight:700;color:var(--dark2);background:var(--gold3);border:1px solid var(--gold);border-radius:4px;padding:2px 6px;white-space:nowrap}}
.sg4{{display:grid;grid-template-columns:1fr 1fr;gap:2px;margin-top:3px}}
.sc{{display:flex;flex-direction:column;align-items:center;gap:0;background:var(--bg2);border-radius:4px;padding:3px 2px;border:1px solid var(--border);transition:all .15s}}
.sl{{font-size:6.5px;font-weight:700;color:var(--text3);text-transform:uppercase;text-align:center;line-height:1.3;letter-spacing:.2px}}
.sv{{font-family:'Playfair Display',serif;font-size:15px;color:var(--dark);line-height:1}}
.str-row{{display:flex;gap:2px;margin-top:2px}}
.sc3{{display:flex;flex-direction:column;align-items:center;gap:0;background:var(--bg2);border-radius:4px;padding:3px 2px;flex:1;border:1px solid var(--border);transition:all .15s}}
.sc3 .sl{{font-size:6px;font-weight:700;color:var(--text3);text-transform:uppercase;text-align:center}}
.sc3 .sv{{font-family:'Playfair Display',serif;font-size:11px;color:var(--dark);line-height:1}}
.bw{{height:2px;background:var(--border);border-radius:1px;overflow:hidden;margin-top:3px}}
.bf{{height:100%;border-radius:1px}}
.sg1{{color:#2d6a4f;font-weight:800}}.sg2{{color:#7a5c00;font-weight:800}}.sg3{{color:#9b2020;font-weight:800}}
.noimg{{display:flex;flex-direction:column;align-items:center;gap:4px;color:var(--border2);font-size:8px}}

/* WMS STOK SATIRI */
.wms-row{{display:grid;grid-template-columns:repeat(3,1fr);gap:3px;margin-top:2px;padding-top:3px;border-top:1px dashed var(--border2)}}
.stok-row.sg2{{grid-template-columns:1fr 1fr!important}}
.wms-v{{color:#4BACC8!important;font-weight:700}}
.depo-v{{color:#7BA7D8!important;font-weight:700}}
.tot-stok-v{{color:#9aD4E8!important;font-weight:700}}

/* ARAMA - Excel Filtre Stili */
.search-bar{{padding:.5rem .8rem .3rem;border-bottom:1px solid var(--border);display:flex;flex-direction:column;gap:.4rem;background:var(--bg)}}
.srch-wrap{{display:flex;align-items:center;gap:.4rem;background:var(--card);border:1.5px solid var(--border2);border-radius:8px;padding:.3rem .6rem;transition:border-color .2s}}
.srch-wrap:focus-within{{border-color:var(--dark2);box-shadow:0 0 0 2px rgba(200,168,75,.15)}}
.srch-icon{{font-size:13px;color:var(--text3);flex-shrink:0}}
.srch-inp{{flex:1;border:none;outline:none;background:transparent;font-family:inherit;font-size:11px;color:var(--text);padding:2px 0}}
.srch-inp::placeholder{{color:var(--text3)}}
.srch-clear{{background:none;border:none;color:var(--text3);cursor:pointer;font-size:14px;padding:0 2px;flex-shrink:0;line-height:1}}
.srch-clear:hover{{color:var(--dark)}}
/* Chips */
.sel-chips{{display:flex;flex-wrap:wrap;gap:4px;min-height:0}}
.sel-chip{{display:flex;align-items:center;gap:5px;background:var(--dark2);color:#f0d8a0;border:1px solid var(--gold);border-radius:12px;padding:2px 8px 2px 10px;font-size:9px;font-weight:700;cursor:pointer;transition:all .15s}}
.sel-chip:hover{{background:#5a1010;border-color:#ff6666}}
.sel-chip-clear{{background:transparent;color:var(--text3);border-color:var(--border2)}}
.sel-chip-clear:hover{{background:var(--bg3);color:var(--dark)}}
.sel-chip .ch-x{{font-size:11px;opacity:.6}}
/* Dropdown */
.srch-wrap-rel{{position:relative}}
.srch-results{{
  position:absolute;top:calc(100% + 3px);left:0;right:0;z-index:500;
  background:var(--card);border:1.5px solid var(--border2);
  border-radius:8px;max-height:240px;overflow-y:auto;
  box-shadow:0 6px 24px rgba(0,0,0,.15);
}}
.srch-item{{padding:6px 10px;font-size:10px;cursor:pointer;display:flex;align-items:center;gap:7px;border-bottom:1px solid var(--border)}}
.srch-item:last-child{{border-bottom:none}}
.srch-item:hover{{background:var(--bg2)}}
.srch-item.sel{{background:var(--gold3)}}
.srch-selall{{background:var(--bg2);font-weight:700;position:sticky;top:0;z-index:1;border-bottom:2px solid var(--border2)!important}}
.srch-selall:hover{{background:var(--bg3)}}
.si-chk-box{{font-size:14px;color:var(--dark2);flex-shrink:0;width:16px;text-align:center}}
.srch-item.sel .si-chk-box{{color:var(--dark)}}
.si-code{{font-family:monospace;font-size:9.5px;font-weight:700;color:var(--dark2);flex-shrink:0}}
.si-info{{font-size:9px;color:var(--text3);flex:1}}
.si-cnt{{margin-left:auto;font-size:9px;color:var(--text3);font-weight:400}}

/* ALT TOPLAM */
.tot-bar{{display:flex;flex-wrap:wrap;gap:3px;padding:.4rem 0 .3rem;margin-bottom:.4rem;border-top:2px solid var(--border);border-bottom:1px solid var(--border)}}
.tot-cell{{display:flex;flex-direction:column;align-items:center;background:var(--dark);border-radius:5px;padding:3px 7px;min-width:60px}}
.tot-cell.tot-lbl{{background:var(--dark2);min-width:80px;align-items:flex-start}}
.tot-l{{font-size:6.5px;font-weight:700;color:#8a7a5a;text-transform:uppercase;letter-spacing:.5px}}
.tot-v{{font-family:'Playfair Display',serif;font-size:14px;color:#f0d8a0;line-height:1;font-weight:700}}
.tot-n{{font-size:9px;color:var(--gold);font-weight:600}}
.tot-lbl>span:first-child{{font-size:10px;font-weight:700;color:#f0d8a0;letter-spacing:1px;text-transform:uppercase}}

/* SIFIR TOGGLE */
.zero-btn{{display:flex;align-items:center;gap:6px;margin:.3rem 0 .5rem;padding:5px 12px;border-radius:20px;border:1px dashed var(--border2);background:var(--bg2);color:var(--text2);font-size:10px;font-weight:600;cursor:pointer;font-family:'DM Sans',sans-serif;transition:all .15s}}
.zero-btn:hover{{border-color:var(--dark);color:var(--dark)}}
.zero-btn.zero-open{{background:var(--dark2);color:#f0d8a0;border-color:var(--dark2)}}
.zarr{{font-size:8px;transition:transform .2s}}
.zbadge{{font-size:9px;padding:1px 6px;border-radius:10px;background:rgba(200,168,75,.15);color:var(--gold2)}}
.stok-row{{margin-top:2px;display:grid;grid-template-columns:1fr 1fr 1fr;gap:2px}}
.sc-stok{{background:var(--bg3)!important}}
.sc-stok .sl{{color:var(--dark2)!important;font-weight:700!important}}
.sc-stok .sv{{font-size:13px!important}}

/* STR PANEL */
.str-nav{{background:var(--bg2);padding:.35rem 1rem;display:flex;gap:5px;border-bottom:1px solid var(--border);overflow-x:auto;-webkit-overflow-scrolling:touch}}
.str-nav::-webkit-scrollbar{{display:none}}
.snav{{font-size:10px;font-weight:600;padding:4px 10px;border-radius:16px;border:1px solid var(--border2);background:var(--bg);color:var(--text2);cursor:pointer;font-family:'DM Sans',sans-serif;transition:all .15s;white-space:nowrap}}
.snav.act{{background:var(--dark);color:var(--gold);border-color:var(--dark)}}
.str-wrap{{padding:1rem 1rem 2rem}}
.str-section{{margin-bottom:1.5rem}}
.str-title{{font-family:'Playfair Display',serif;font-size:.9rem;letter-spacing:1px;color:var(--dark);margin-bottom:.8rem;padding-bottom:.4rem;border-bottom:2px solid var(--gold)}}
.stw{{background:var(--card);border-radius:var(--r2);border:1px solid var(--border);overflow:hidden;overflow-x:auto;-webkit-overflow-scrolling:touch}}
.stw table{{width:100%;border-collapse:collapse;min-width:480px}}
.stw th{{background:var(--dark);color:var(--gold);font-size:9px;font-weight:700;letter-spacing:.7px;text-transform:uppercase;padding:8px 10px;text-align:left;white-space:nowrap}}
.stw td{{padding:7px 10px;border-bottom:1px solid var(--bg2);font-size:11px;font-weight:500;color:var(--dark);white-space:nowrap}}
.stw tr:last-child td{{border-bottom:none}}.stw tr:hover td{{background:var(--bg)}}
.sbi{{display:inline-block;height:3px;border-radius:2px;margin-left:5px;vertical-align:middle;opacity:.6}}
.str-tot{{background:var(--dark)!important}}
.str-tot td{{color:#f0d8a0!important;font-size:11px!important;font-weight:700!important;border-bottom:none!important;padding:8px 10px!important}}
.str-tot td b{{color:#f0d8a0!important}}
.str-tot td .sbi{{opacity:.8}}
.str-tot:hover td{{background:var(--dark2)!important}}

/* STR subpanel */
.str-sub{{display:none}}.str-sub.active{{display:block}}

/* HOSS */
.hoss-kpi{{background:var(--dark2);padding:.6rem 1rem;display:flex;gap:1rem;align-items:center;flex-wrap:wrap;border-bottom:1px solid #4a3c22}}
.hoss-kpi .kpi-l{{color:#8a7a5a}}
.hoss-kpi .kpi-v{{color:#f0d8a0}}

/* LAYOUT */
.panel{{display:none}}.panel.active{{display:block}}
.sub-panel{{display:none}}.sub-panel.active{{display:block}}

/* FOOTER */
.footer{{text-align:center;padding:.7rem 1rem;font-size:10px;color:var(--text3);border-top:1px solid var(--border);background:var(--bg2);line-height:1.8}}
.footer b{{color:var(--text2);font-weight:600}}

/* MOBILE */
@media(max-width:600px){{
  .brand{{font-size:.85rem;letter-spacing:1px}}
  .hb{{display:none}}
  .header{{padding:.6rem .8rem}}
  .kpi-bar{{grid-template-columns:repeat(3,1fr);gap:.25rem;padding:.4rem .6rem}}
  .kpi{{padding:.25rem .3rem}}
  .kpi-v{{font-size:14px}}
  .kpi-l{{font-size:6.5px}}
  .kdiv{{display:none}}
  .tab-bar{{padding:0 .5rem}}
  .tab{{padding:.45rem .6rem;font-size:10px}}
  .sub-tabs{{padding:0 .5rem}}
  .stab{{padding:.4rem .6rem;font-size:10px}}
  .toolbar2{{padding:.35rem .6rem;gap:.3rem;flex-direction:column}}
  .tb-left{{gap:4px}}
  .tb-right{{gap:3px;flex-wrap:wrap}}
  .tlbl2{{font-size:8px}}
  .dd-btn{{padding:4px 7px;font-size:10px;min-height:28px}}
  .sbtn{{font-size:9px;padding:3px 5px;min-height:24px}}
  .gw{{padding:.5rem .6rem 1.5rem}}
  .grid{{grid-template-columns:1fr 1fr;gap:5px}}
  .iw{{height:110px}}
  .cbody{{padding:5px 6px 6px;gap:2px}}
  .cname{{font-size:9.5px}}
  .sv{{font-size:12px}}
  .sc .sl{{font-size:6px}}
  .sg4{{gap:2px}}
  .str-row{{gap:2px}}
  .sc3 .sv{{font-size:10px}}
  .stok-row{{gap:2px}}
  .sc-stok .sv{{font-size:11px!important}}
  .bw{{margin-top:2px}}
  .str-wrap{{padding:.6rem .6rem 1.5rem}}
  .str-nav{{padding:.3rem .6rem}}
  .stw table{{min-width:360px}}
  .stw th,.stw td{{padding:5px 7px;font-size:10px}}
  .tot-bar{{gap:2px;padding:.3rem 0}}
  .tot-cell{{padding:2px 5px;min-width:50px}}
  .tot-v{{font-size:12px}}
  .tot-l{{font-size:6px}}
  .sezon-bar{{padding:.35rem .6rem}}
  .sezon-btn{{padding:5px 10px;font-size:10px}}
  .shdr{{font-size:8px}}
  .zero-btn{{font-size:9px;padding:4px 9px}}
}}

@media(max-width:360px){{
  .grid{{grid-template-columns:1fr 1fr}}
  .kpi-bar{{grid-template-columns:repeat(2,1fr)}}
  .iw{{height:95px}}
  .sv{{font-size:11px}}
}}

@media(min-width:768px){{
  .toolbar2{{flex-direction:row;align-items:center;justify-content:space-between}}
  .grid{{grid-template-columns:repeat(auto-fill,minmax(165px,1fr));gap:9px}}
}}

@media(min-width:1024px){{
  .header{{padding:1rem 2rem}}
  .tab-bar{{padding:0 2rem}}
  .kpi-bar{{display:flex;gap:1rem;grid-template-columns:none;background:var(--bg2);padding:.6rem 2rem}}
  .kpi{{background:transparent;border:none;padding:0}}
  .kdiv{{display:block;width:1px;height:26px;background:var(--border2);flex-shrink:0}}
  .gw{{padding:1rem 2rem 2rem}}
  .str-wrap{{padding:1.2rem 2rem 2rem}}
  .str-nav{{padding:.35rem 2rem}}
  .sub-tabs{{padding:0 2rem}}
  .grid{{grid-template-columns:repeat(auto-fill,minmax(175px,1fr));gap:10px}}
  .iw{{height:160px}}
  .brand{{font-size:1.3rem}}
  .tot-bar{{gap:5px}}
  .tot-cell{{min-width:70px}}
  .tot-v{{font-size:16px}}
}}
</style></head><body>
<div class="header">
  <div class="logo-wrap">
    <div class="logo-icon">🐊</div>
    <div>
      <div class="brand">Lacoste 26AW Takip Raporu</div>
      <div class="brand-sub">{hafta_str} &nbsp;·&nbsp; {excel_adi}</div>
    </div>
  </div>
  <div class="hb">{total} Option &nbsp;·&nbsp; ANA TABLO</div>
</div>
<div class="tab-bar">{tabs}</div>
{panels}
{hoss_panel}
<div class="footer">Excel: <b>{excel_adi}</b> · STR = (Mag+Online Satis) / Siparis x 100 · <b>{hafta_str}</b></div>
<script>
{js_data}
{js_analiz}
var HOSS_URUNLER={hoss_js_urunler};
var HOSS_KAT={hoss_js_kat};
var HOSS_UST={hoss_js_ust};
var HOSS_CIN={hoss_js_cin};
var KMETA={kmeta};
var STATE={{{",".join([f'"{k}":{{f:"all",kat:"all",aile:[],s:"top_ytd",d:"desc"}}' for k in KANALLAR])}}};
var HOSS_STATE={{f:"all",sort_k:"SSTEP"}};
var gL={{MEN:"Erkek",WOMEN:"Kadin",CHILDREN:"Cocuk",ACCESSORIES:"Aksesuar"}};
function sc(v){{return v>=20?"sg1":v>=10?"sg2":"sg3";}}

function renderGrid(k){{
  const grid=document.getElementById("grid-"+k);
  const st=STATE[k];
  const data=window["URUNLER_"+k]||[];
  let vis=[...data];
  if(st.f!=="all") vis=vis.filter(p=>p.gender===st.f);
  if(st.kat&&st.kat!=="all") vis=vis.filter(p=>p.ana_kat===st.kat);
  if(st.aile&&st.aile.length>0) vis=vis.filter(p=>st.aile.includes(p.aile));
  const asc=st.d==="asc";
  vis.sort((a,b)=>asc?a[st.s]-b[st.s]:b[st.s]-a[st.s]);
  const renk=KMETA[k].renk;
  const maxV=Math.max(...vis.map(p=>Math.max(0,p[st.s])),1);
  grid.innerHTML="";
  vis.forEach((p,i)=>{{
    const isTop=!asc&&i<3;
    const card=document.createElement("div");
    card.className="card"+(isTop?" top":"");
    if(isTop) card.style.borderColor=renk;
    const rnkLbl=i===0&&!asc?"#1 COK SATAN":"#"+(i+1);
    const rnkGld=isTop?"gld":"";
    const bp=Math.round(Math.max(0,p[st.s])/maxV*100);
    const wid="iw_"+k+"_"+p.code.replace(/[^a-zA-Z0-9]/g,"_");
    const hiBg="background:"+renk+"22;border-color:"+renk;
    var hiSt=function(active){{return active?'style="'+hiBg+'"':'';}};
    var hiSv=function(active){{return active?'style="color:'+renk+';font-size:18px"':'';}};
    var s1=st.s==="top_ytd", s2=st.s==="mag_ytd",
        s3=st.s==="top_sh"||st.s==="mag_sh",
        s4=st.s==="onl_ytd"||st.s==="onl_sh",
        s5=st.s==="str_h", s6=st.s==="str_d", s7=st.s==="ros",
        s8=st.s==="stok",
        s9=st.s==="mag_mtd"||st.s==="onl_mtd"||st.s==="top_mtd";
    var storyHtml=p.story&&p.story!=="0"&&p.story!==""?'<div class="cstory">'+p.story+'</div>':"";
    card.innerHTML=
      '<div class="rnk '+rnkGld+'">'+rnkLbl+'</div>'+
      '<div class="ctag">'+p.ust_kat+'</div>'+
      '<div class="iw" id="'+wid+'"><div class="spin"></div></div>'+
      '<div class="cbody">'+
        '<div class="crow">'+
          '<span class="cgender">'+(gL[p.gender]||p.gender)+'</span>'+
          '<div class="cright">'+
            '<div class="ckat">'+(p.alt_aile&&p.alt_aile!=="0"?p.alt_aile:p.alt_kat)+'</div>'+
            '<div class="calt">'+p.ana_kat+'</div>'+
          '</div>'+
        '</div>'+
        '<div class="cname">'+(p.aile&&p.aile!=="0"?p.aile:p.code.split(".")[0])+'</div>'+
        storyHtml+
        '<div class="ccode">'+p.code+'</div>'+
        '<div class="sg4">'+
          '<div class="sc" '+hiSt(s1)+'><span class="sl">Top YTD</span><span class="sv" '+hiSv(s1)+'>'+p.top_ytd+'</span></div>'+
          '<div class="sc" '+hiSt(s2)+'><span class="sl">Mag YTD</span><span class="sv" '+hiSv(s2)+'>'+p.mag_ytd+'</span></div>'+
          '<div class="sc" '+hiSt(s3)+'><span class="sl">Son Hafta</span><span class="sv" '+hiSv(s3)+'>'+p.top_sh+'</span></div>'+
          '<div class="sc" '+hiSt(s4)+'><span class="sl">Online YTD</span><span class="sv" '+hiSv(s4)+'>'+p.onl_ytd+'</span></div>'+
        '</div>'+
        '<div class="stok-row">'+
          '<div class="sc sc-stok" '+hiSt(s8)+'><span class="sl">Mag Stok</span><span class="sv" '+hiSv(s8)+'>'+p.stok+'</span></div>'+
          '<div class="sc sc-stok" '+hiSt(s9&&st.s==="mag_mtd")+'><span class="sl">MTD Mag</span><span class="sv" '+hiSv(s9&&st.s==="mag_mtd")+'>'+p.mag_mtd+'</span></div>'+
          '<div class="sc sc-stok" '+hiSt(s9&&st.s==="onl_mtd")+'><span class="sl">MTD Online</span><span class="sv" '+hiSv(s9&&st.s==="onl_mtd")+'>'+p.onl_mtd+'</span></div>'+
        '</div>'+
        '<div class="str-row">'+
          '<div class="sc3" '+hiSt(s5)+'><span class="sl">STR D.H</span><span class="sv '+sc(p.str_h)+'" '+hiSv(s5)+'>'+p.str_h+'%</span></div>'+
          '<div class="sc3" '+hiSt(s6)+'><span class="sl">STR D.D</span><span class="sv '+sc(p.str_d)+'" '+hiSv(s6)+'>'+p.str_d+'%</span></div>'+
          '<div class="sc3" '+hiSt(s7)+'><span class="sl">ROS</span><span class="sv" '+hiSv(s7)+'>'+p.ros+'</span></div>'+
        '</div>'+
        '<div class="bw"><div class="bf" style="width:'+bp+'%;background:'+renk+'"></div></div>'+
      '</div>';
    grid.appendChild(card);
    (function(src,code,wid){{
      if(!src){{ noImg(wid,code); return; }}
      var im=new Image();
      im.onload=function(){{
        var w=document.getElementById(wid);
        if(w){{
          im.style.cssText="width:100%;height:100%;object-fit:contain;padding:8px;cursor:zoom-in";
          im.onclick=function(){{window.open(src,"_blank");}};
          w.innerHTML=""; w.appendChild(im);
        }}
      }};
      im.onerror=function(){{noImg(wid,code);}};
      im.src=src;
    }})(p.img,p.code,wid);
  }});
  var lblMap={{
    top_ytd:"Tum Satislar",mag_ytd:"Magaza YTD",onl_ytd:"Online YTD",
    top_sh:"Son Hafta Toplam",mag_sh:"Son Hafta Magaza",
    onl_sh:"Son Hafta Online",str_h:"STR",str_d:"STR (Devir Dahil)",ros:"ROS",
    stok:"Magaza Stok",
    mag_mtd:"MTD Magaza",onl_mtd:"MTD Online",top_mtd:"MTD Toplam",
  }};
  document.getElementById("lbl-"+k).textContent=(asc?"En Az Satan":"En Cok Satan")+" — "+(lblMap[st.s]||st.s);
  document.getElementById("cnt-"+k).textContent=vis.length+" urun";
}}
function renderStr(k){{
  const a=window["ANALIZ_"+k]||{{}};
  const kat=(a.kat)||[];
  const renk=KMETA[k].renk;
  let rows="";
  kat.forEach(d=>{{
    const bw=Math.min((d.str_h||0)*2,100);
    rows+=`<tr>
      <td style="font-weight:700">${{d.kat}}</td>
      <td>${{(d.sip_h||0).toLocaleString("en")}}</td><td>${{(d.sip_d||0).toLocaleString("en")}}</td>
      <td>${{d.mag_ytd||0}}</td><td>${{d.onl_ytd||0}}</td><td>${{d.top_ytd||0}}</td>
      <td><b class="${{sc(d.str_h||0)}}">${{d.str_h||0}}%</b>
        <div class="sbi" style="width:${{bw}}px;background:${{renk}}"></div></td>
      <td><b class="${{sc(d.str_d||0)}}">${{d.str_d||0}}%</b></td>
    </tr>`;
  }});
  document.getElementById("st-"+k).innerHTML=
    `<div class="stw"><table><thead><tr>
      <th>Ana Kategori</th><th>Sip D.Haric</th><th>Sip D.Dahil</th>
      <th>Mag YTD</th><th>Online YTD</th><th>Toplam YTD</th>
      <th>STR D.Haric</th><th>STR D.Dahil</th>
    </tr></thead><tbody>${{rows}}</tbody></table></div>`;

  // Cinsiyet STR tablosu
  const cinData=(window["ANALIZ_"+k]||{{}}).cin||[];
  let cinRows="";
  cinData.forEach(d=>{{
    const b=Math.min((d.str_h||0)*2,100);
    cinRows+=`<tr>
      <td style="font-weight:700">${{gL[d.cin]||d.cin}}</td>
      <td>${{(d.sip_h||0).toLocaleString("en")}}</td><td>${{(d.sip_d||0).toLocaleString("en")}}</td>
      <td>${{d.mag_ytd||0}}</td><td>${{d.onl_ytd||0}}</td><td>${{d.top_ytd||0}}</td>
      <td><b class="${{sc(d.str_h||0)}}">${{d.str_h||0}}%</b>
        <div class="sbi" style="width:${{b}}px;background:${{renk}}"></div></td>
      <td><b class="${{sc(d.str_d||0)}}">${{d.str_d||0}}%</b></td>
    </tr>`;
  }});
  const cinEl=document.getElementById("st-cin-"+k);
  if(cinEl) cinEl.innerHTML=
    `<div class="stw"><table><thead><tr>
      <th>Cinsiyet</th><th>Sip D.Haric</th><th>Sip D.Dahil</th>
      <th>Mag YTD</th><th>Online YTD</th><th>Toplam YTD</th>
      <th>STR D.Haric</th><th>STR D.Dahil</th>
    </tr></thead><tbody>${{cinRows}}</tbody></table></div>`;

  // Aile STR tablosu
  const aileEl=document.getElementById("st-aile-"+k);
  if(aileEl){{
    const aileData=(window["ANALIZ_"+k]||{{}}).aile||[];
    let aileRows="";
    aileData.forEach(d=>{{
      const b=Math.min((d.str_h||0)*2.5,100);
      aileRows+=`<tr>
        <td style="font-weight:700">${{d.aile}}</td>
        <td>${{(d.sip_h||0).toLocaleString("en")}}</td>
        <td>${{d.mag_ytd||0}}</td><td>${{d.onl_ytd||0}}</td><td>${{d.top_ytd||0}}</td>
        <td>${{d.stok||0}}</td>
        <td><b class="${{sc(d.str_h||0)}}">${{d.str_h||0}}%</b>
          <div class="sbi" style="width:${{b}}px;background:${{renk}}"></div></td>
        <td><b class="${{sc(d.str_d||0)}}">${{d.str_d||0}}%</b></td>
      </tr>`;
    }});
    aileEl.innerHTML=
      `<div class="stw"><table><thead><tr>
        <th>Ayakkabi Aile</th><th>Siparis</th>
        <th>Mag YTD</th><th>Online YTD</th><th>Toplam YTD</th><th>Stok</th>
        <th>STR D.Haric</th><th>STR D.Dahil</th>
      </tr></thead><tbody>${{aileRows}}</tbody></table></div>`;
  }}
}}

function updateKPI(k){{
  const a=((window["ANALIZ_"+k]||{{}}).toplam)||{{}};
  document.getElementById("kv1-"+k).textContent=(a.sip_h||0).toLocaleString("en");
  document.getElementById("kv2-"+k).textContent=(a.sip_d||0).toLocaleString("en");
  document.getElementById("kv3-"+k).textContent=(a.mag_ytd||0).toLocaleString("en");
  document.getElementById("kv4-"+k).textContent=(a.onl_ytd||0).toLocaleString("en");
  document.getElementById("kv5-"+k).textContent=(a.top_ytd||0).toLocaleString("en");
  document.getElementById("kv6-"+k).textContent=(a.str_h||0)+"%";
  document.getElementById("kv7-"+k).textContent=(a.str_d||0)+"%";
  document.getElementById("kv8-"+k).textContent=(a.stok||0).toLocaleString("en");
}}

function noImg(wid,code){{
  const w=document.getElementById(wid);if(!w)return;
  w.innerHTML='<div class="noimg"><svg width="26" height="26" viewBox="0 0 26 26" fill="none"><rect x="1" y="1" width="24" height="24" rx="3" stroke="#d4c9b0" stroke-width="1.5"/><path d="M3 19l4-6 4 4 4-5 4 7H3z" stroke="#d4c9b0" stroke-width="1.5" fill="none"/></svg>'+code.split(".")[0]+'</div>';
}}

function indir(k){{
  if(typeof sendPrompt==="function"){{
    sendPrompt(k+" gitmeyen Excel indir");
  }}else{{
    alert("Komut satirinda calistirin:\\npython lacoste_rapor_v2.py --excel DOSYA.xlsx --indir");
  }}
}}

function renderHOSS(){{
  const grid=document.getElementById("grid-HOSS");if(!grid)return;
  const sk=HOSS_STATE.sort_k;
  let vis=[...HOSS_URUNLER];
  if(HOSS_STATE.f!=="all") vis=vis.filter(p=>p.gender===HOSS_STATE.f);
  grid.innerHTML="";
  vis.forEach((p,i)=>{{
    const card=document.createElement("div");
    card.className="card";
    card.style.borderTop="3px solid #7b3f00";
    const satis=p[sk+"_top_ytd"]||0;
    const stok =p[sk+"_stok"]||0;
    const str  =p[sk+"_str_h"]||0;
    const sip  =p[sk+"_sip_h"]||0;
    const wid="iw_HOSS_"+p.code.replace(/[^a-zA-Z0-9]/g,"_");
    card.innerHTML=`
      <div class="rnk">#${{i+1}}</div>
      <div class="ctag">${{p.ust_kat}}</div>
      <div class="iw" id="${{wid}}"><div class="spin"></div></div>
      <div class="cbody">
        <div class="crow"><span class="cgender">${{gL[p.gender]||p.gender}}</span>
          <div class="cright"><div class="ckat">${{p.ana_kat}}</div></div></div>
        <div class="cname">${{p.ust_kat}}</div>
        <div class="cstory">${{p.story}}</div>
        <div class="ccode">${{p.code}}</div>
        <div class="sg4">
          <div class="sc"><span class="sl">Siparis</span><span class="sv">${{sip}}</span></div>
          <div class="sc hi"><span class="sl">Satis (${{sk}})</span><span class="sv">${{satis}}</span></div>
          <div class="sc"><span class="sl">Stok</span><span class="sv">${{stok}}</span></div>
          <div class="sc"><span class="sl">STR</span><span class="sv ${{sc(str)}}">${{str}}%</span></div>
        </div>
      </div>`;
    grid.appendChild(card);
    (function(src,code,elId){{
      if(!src){{noImg(elId,code);return;}}
      var im=new Image();
      im.onload=function(){{var w=document.getElementById(elId);if(w){{
        im.style.cssText="width:100%;height:100%;object-fit:contain;padding:8px;cursor:zoom-in";
        im.onclick=function(){{window.open(src,"_blank");}};
        w.innerHTML="";w.appendChild(im);
      }}}};
      im.onerror=function(){{noImg(elId,code);}};
      im.src=src;
    }})(p.img,p.code,wid);
  }});
  document.getElementById("lbl-HOSS").textContent=sk+" Satisina Gore";
  document.getElementById("cnt-HOSS").textContent=vis.length+" urun";
}}

function renderHOSSStr(){{
  const KANALLER=Object.keys(KMETA);
  function hossTable(data, labelKey){{
    let rows="";
    data.forEach(d=>{{
      rows+="<tr><td style='font-weight:700'>"+d[labelKey]+"</td>";
      KANALLER.forEach(k=>{{
        const sh=d[k+"_sip_h"]||0;
        const ty=d[k+"_top_ytd"]||0;
        const str=d[k+"_str_h"]||0;
        rows+=`<td>${{sh}}</td><td>${{ty}}</td><td><b class="${{sc(str)}}">${{str}}%</b></td>`;
      }});
      rows+="</tr>";
    }});
    let hdr="<tr><th>"+labelKey+"</th>";
    KANALLER.forEach(k=>hdr+=`<th>${{k}} Sip</th><th>${{k}} Satis</th><th>${{k}} STR</th>`);
    hdr+="</tr>";
    return `<div class="stw"><table><thead>${{hdr}}</thead><tbody>${{rows}}</tbody></table></div>`;
  }}
  const katEl=document.getElementById("st-HOSS-kat");
  if(katEl) katEl.innerHTML=hossTable(HOSS_KAT,"kat");
  const ustEl=document.getElementById("st-HOSS-ust");
  if(ustEl) ustEl.innerHTML=hossTable(HOSS_UST,"ust");

  // Kod bazli STR
  const kodEl=document.getElementById("st-HOSS-kod");
  if(kodEl){{
    let rows="";
    HOSS_URUNLER.forEach(p=>{{
      rows+="<tr><td style='font-family:monospace;font-weight:600'>"+p.code+"</td><td>"+p.gender+"</td><td>"+p.ana_kat+"</td><td>"+p.ust_kat+"</td>";
      KANALLER.forEach(k=>{{
        const sh=p[k+"_sip_h"]||0;
        const ty=p[k+"_top_ytd"]||0;
        const st=p[k+"_stok"]||0;
        const str=p[k+"_str_h"]||0;
        rows+=`<td>${{sh}}</td><td>${{ty}}</td><td>${{st}}</td><td><b class="${{sc(str)}}">${{str}}%</b></td>`;
      }});
      rows+="</tr>";
    }});
    let hdr="<tr><th>Kod</th><th>Cinsiyet</th><th>Ana Kat</th><th>Ust Kat</th>";
    KANALLER.forEach(k=>hdr+=`<th>${{k}} Sip</th><th>${{k}} Satis</th><th>${{k}} Stok</th><th>${{k}} STR</th>`);
    hdr+="</tr>";
    kodEl.innerHTML=`<div class="stw"><table><thead>${{hdr}}</thead><tbody>${{rows}}</tbody></table></div>`;
  }}
}}

function renderHOSSOzet(){{
  const cinEl=document.getElementById("st-HOSS-cin");
  if(!cinEl)return;
  const KANALLER=Object.keys(KMETA);
  let rows="";
  HOSS_CIN.forEach(d=>{{
    rows+="<tr><td style='font-weight:700'>"+(gL[d.cin]||d.cin)+"</td>";
    KANALLER.forEach(k=>{{
      const sh=d[k+"_sip_h"]||0;
      const ty=d[k+"_top_ytd"]||0;
      const st=d[k+"_stok"]||0;
      const str=d[k+"_str_h"]||0;
      rows+=`<td>${{sh}}</td><td>${{ty}}</td><td>${{st}}</td><td><b class="${{sc(str)}}">${{str}}%</b></td>`;
    }});
    rows+="</tr>";
  }});
  let hdr="<tr><th>Cinsiyet</th>";
  KANALLER.forEach(k=>hdr+=`<th>${{k}} Sip</th><th>${{k}} Satis</th><th>${{k}} Stok</th><th>${{k}} STR</th>`);
  hdr+="</tr>";
  cinEl.innerHTML=`<div class="stw"><table><thead>${{hdr}}</thead><tbody>${{rows}}</tbody></table></div>`;
}}

{JS_DD_FUNCS}
function toggleAile(k){{
  const dd=document.getElementById("aile-dd-"+k);
  if(dd.style.display==="none"){{
    dd.style.display="block";
    const data=window["URUNLER_"+k]||[];
    const aileler=[...new Set(data.map(p=>p.aile).filter(a=>a&&a!=="0"&&a!=="-"))].sort();
    const list=document.getElementById("aile-list-"+k);
    if(!list.children.length){{
      list.innerHTML=aileler.map(a=>
        `<label style="display:flex;align-items:center;gap:6px;padding:3px 0;cursor:pointer;font-size:11px;color:var(--dark)">
          <input type="checkbox" value="${{a}}" onchange="aileChange('${k}')"> ${{a}}
        </label>`
      ).join("");
    }}
    document.addEventListener("click",function outsideClick(e){{
      if(!document.getElementById("faile-"+k).contains(e.target)){{
        dd.style.display="none";
        document.removeEventListener("click",outsideClick);
      }}
    }},{{capture:true}});
  }} else dd.style.display="none";
}}
function aileChange(k){{
  const checks=document.querySelectorAll("#aile-list-"+k+" input:checked");
  STATE[k].aile=[...checks].map(c=>c.value);
  const val=document.getElementById("ddv-aile-"+k);
  if(val) val.textContent=STATE[k].aile.length?(STATE[k].aile.length+" secili"):"Tumu";
  const btn=document.getElementById("ddb-aile-"+k);
  if(btn){{STATE[k].aile.length?btn.classList.add("active"):btn.classList.remove("active");}}
  renderGrid(k);
}}
function aileTumunu(k){{
  document.querySelectorAll("#aile-list-"+k+" input").forEach(c=>c.checked=true);
  aileChange(k);
}}
function aileTemizle(k){{
  document.querySelectorAll("#aile-list-"+k+" input").forEach(c=>c.checked=false);
  STATE[k].aile=[];
  const btn=document.getElementById("aile-toggle-"+k);
  btn.textContent="Tum Aileler ▾";
  renderGrid(k);
}}

function switchTab(k,btn){{
  document.querySelectorAll(".tab").forEach(t=>t.classList.remove("active"));
  document.querySelectorAll(".panel").forEach(p=>p.classList.remove("active"));
  btn.classList.add("active");
  document.getElementById("panel-"+k).classList.add("active");
  if(k==="HOSS"){{
    if(!inited["HOSS"]){{
      inited["HOSS"]=true;
      // HOSS kanal secim toolbar
      const sg=document.getElementById("sg-HOSS");
      if(sg)sg.addEventListener("click",e=>{{
        const b=e.target.closest("[data-k]");if(!b)return;
        sg.querySelectorAll("[data-k]").forEach(x=>x.classList.remove("as"));
        b.classList.add("as");HOSS_STATE.sort_k=b.dataset.k;renderHOSS();
      }});
      // HOSS filtre
      const fg=document.getElementById("fg-HOSS");
      if(fg)fg.addEventListener("click",e=>{{
        const b=e.target.closest("[data-f]");if(!b)return;
        fg.querySelectorAll("[data-f]").forEach(x=>x.classList.remove("af"));
        b.classList.add("af");HOSS_STATE.f=b.dataset.f;renderHOSS();
      }});
    }}
    renderHOSS();
  }} else initK(k);
}}

function toggleAileMenu(k){{
  const m=document.getElementById("aile-menu-"+k);
  if(m) m.classList.toggle("open");
  document.addEventListener("click",function cls(e){{
    if(!e.target.closest("#faile-"+k)){{m&&m.classList.remove("open");document.removeEventListener("click",cls)}}
  }},{{once:false}});
}}

function buildAileMenu(k){{
  const menu=document.getElementById("aile-menu-"+k);
  if(!menu||menu.dataset.built) return;
  menu.dataset.built=1;
  const data=window["URUNLER_"+k]||[];
  const ailer=[...new Set(data.map(p=>p.aile).filter(a=>a&&a!=="0"&&a!=="-"))].sort();
  menu.innerHTML=ailer.map(a=>`
    <div class="aile-item" onclick="toggleAile('${{k}}','${{a}}',this)">
      <input type="checkbox" id="ac-${{k}}-${{a.replace(/[^a-zA-Z0-9]/g,'_')}}" onchange="toggleAile('${{k}}','${{a}}',this.closest('.aile-item'))">
      <label style="cursor:pointer">${{a}}</label>
    </div>`).join("");
  const wrap=document.getElementById("faile-"+k);
  if(wrap){{
    const topBtn=wrap.querySelector("button");
    if(topBtn)topBtn.onclick=()=>{{menu.classList.toggle("open");buildAileMenu(k);}};
  }}
}}

function toggleAile(k,aile,el){{
  const st=STATE[k];
  const idx=st.aile.indexOf(aile);
  const cb=el.querySelector("input[type=checkbox]");
  if(idx>=0){{st.aile.splice(idx,1);el.classList.remove("sel");if(cb)cb.checked=false;}}
  else{{st.aile.push(aile);el.classList.add("sel");if(cb)cb.checked=true;}}
  const wrap=document.getElementById("faile-"+k);
  if(wrap){{
    const topBtn=wrap.querySelector("button");
    if(topBtn){{
      topBtn.textContent=st.aile.length>0?st.aile.length+" secili ":"Tumu ";
      topBtn.innerHTML+="&#x25BE;";
      topBtn.className=st.aile.length>0?"aile-badge":"btn";
    }}
  }}
  renderGrid(k);
}}

function switchStrView(k,view,btn){{
  btn.closest(".str-wrap").querySelectorAll(".str-nav-btn").forEach(b=>b.classList.remove("active"));
  btn.classList.add("active");
  btn.closest(".str-wrap").querySelectorAll(".str-sub").forEach(s=>s.classList.remove("active"));
  const el=document.getElementById("str-"+view+"-view-"+k);
  if(el)el.classList.add("active");
  if(view==="cin")renderStrCin(k);
  else if(view==="aile")renderStrAile(k);
  else renderStr(k);
}}

function renderStrAile(k){{
  const a=window["ANALIZ_"+k]||{{}};
  const aile=(a.aile)||[];
  const renk=KMETA[k].renk;
  const el=document.getElementById("st-aile-"+k);
  if(!el) return;
  let rows="";
  aile.forEach(d=>{{
    const b=Math.min((d.str_h||0)*3,100);
    rows+=`<tr>
      <td style="font-weight:700">${{d.aile}}</td>
      <td>${{(d.sip_h||0).toLocaleString("en")}}</td>
      <td>${{d.mag_ytd||0}}</td><td>${{d.onl_ytd||0}}</td><td>${{d.top_ytd||0}}</td>
      <td>${{d.stok||0}}</td>
      <td><b class="${{sc(d.str_h||0)}}">${{d.str_h||0}}%</b>
        <div class="sbi" style="width:${{b}}px;background:${{renk}}"></div></td>
      <td><b class="${{sc(d.str_d||0)}}">${{d.str_d||0}}%</b></td>
    </tr>`;
  }});
  el.innerHTML=`<div class="stw"><table><thead><tr>
    <th>Urun Ailesi</th><th>Sip D.Haric</th><th>Mag YTD</th><th>Online YTD</th>
    <th>Toplam YTD</th><th>Stok</th><th>STR D.Haric</th><th>STR D.Dahil</th>
  </tr></thead><tbody>${{rows}}</tbody></table></div>`;
}}

function renderStrCin(k){{
  const a=window["ANALIZ_"+k]||{{}};
  const cin=(a.cin)||[];
  const renk=KMETA[k].renk;
  const el=document.getElementById("st-cin-"+k);
  if(!el) return;
  let rows="";
  cin.forEach(d=>{{
    const b=Math.min((d.str_h||0)*3,100);
    rows+=`<tr>
      <td style="font-weight:700">${{gL[d.cin]||d.cin}}</td>
      <td>${{(d.sip_h||0).toLocaleString("en")}}</td><td>${{(d.sip_d||0).toLocaleString("en")}}</td>
      <td>${{d.mag_ytd||0}}</td><td>${{d.onl_ytd||0}}</td><td>${{d.top_ytd||0}}</td>
      <td><b class="${{sc(d.str_h||0)}}">${{d.str_h||0}}%</b>
        <div class="sbi" style="width:${{b}}px;background:${{renk}}"></div></td>
      <td><b class="${{sc(d.str_d||0)}}">${{d.str_d||0}}%</b></td>
    </tr>`;
  }});
  el.innerHTML=`<div class="stw"><table><thead><tr>
    <th>Cinsiyet</th><th>Sip D.Haric</th><th>Sip D.Dahil</th>
    <th>Mag YTD</th><th>Online YTD</th><th>Toplam YTD</th>
    <th>STR D.Haric</th><th>STR D.Dahil</th>
  </tr></thead><tbody>${{rows}}</tbody></table></div>`;
}}

function switchSub(k,s,btn){{
  btn.closest(".sub-tabs").querySelectorAll(".stab").forEach(x=>x.classList.remove("active"));
  btn.classList.add("active");
  document.getElementById("panel-"+k).querySelectorAll(".sub-panel").forEach(p=>p.classList.remove("active"));
  document.getElementById("sp-"+k+"-"+s).classList.add("active");
  if(k==="HOSS"){{
    if(s==="str") renderHOSSStr();
    else if(s==="ozet") renderHOSSOzet();
    else renderHOSS();
  }} else if(s==="s") renderStr(k);
}}

var inited={{}};
document.addEventListener("click",function(e){{
  if(!e.target.closest(".dd-wrap")&&!e.target.closest(".aile-wrap")){{
    document.querySelectorAll(".dd-menu.open").forEach(function(m){{m.classList.remove("open");}});
    document.querySelectorAll(".dd-btn.active").forEach(function(b){{b.classList.remove("active");}});
  }}
}});
function initK(k){{
  if(inited[k])return;inited[k]=true;
  updateKPI(k);setupTB(k);renderGrid(k);
}}
function setupTB(k){{
  // Siralama butonlari
  const sgEl=document.getElementById("sg-"+k);
  if(sgEl) sgEl.addEventListener("click",e=>{{
    const b=e.target.closest("[data-s]");if(!b)return;
    sgEl.querySelectorAll("[data-s]").forEach(x=>x.classList.remove("as"));
    b.classList.add("as");STATE[k].s=b.dataset.s;STATE[k].d=b.dataset.d||"desc";renderGrid(k);
  }});
  // Aile listesi olusturulunca checkbox event baglama
  const aileListEl=document.getElementById("aile-list-"+k);
  if(aileListEl) aileListEl.addEventListener("change",function(){{aileChange(k);}});
}}
document.querySelector(".tab")&&document.querySelector(".tab").click();
</script></body></html>"""


def secim_yap(xlsx_list, prompt):
    print(f"\n{prompt}")
    for i, p in enumerate(xlsx_list):
        print(f"  [{i+1}] {p.name}")
    print("  [0] Cikis")
    while True:
        try:
            s = input(f"Secim (1-{len(xlsx_list)}, Enter={xlsx_list[0].name}): ").strip()
            if s == "": return xlsx_list[0]
            if s == "0": print("Cikiliyor."); sys.exit(0)
            if s.isdigit() and 1 <= int(s) <= len(xlsx_list): return xlsx_list[int(s)-1]
            print(f"  Gecersiz. 1-{len(xlsx_list)} arasinda sayi girin.")
        except (KeyboardInterrupt, EOFError):
            print("\nIptal."); sys.exit(0)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--indir",  action="store_true")
    parser.add_argument("--no-ac",  action="store_true")
    args = parser.parse_args()

    print(f"\n{'='*55}")
    print(f"  Lacoste Sezon Takip Raporu")
    print(f"  {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    print(f"{'='*55}\n")

    xlsx_list = sorted([p for p in SCRIPT_DIR.glob("*.xlsx") if not p.name.startswith("~$")],
                       key=lambda p: p.stat().st_mtime, reverse=True)
    if not xlsx_list:
        log("Klasorde xlsx bulunamadi!", "X"); sys.exit(1)

    aw_yolu = secim_yap(xlsx_list, "26AW dosyasini secin:")
    log(f"26AW: {aw_yolu.name}", "OK")
    ss_yolu = secim_yap(xlsx_list, "26SS dosyasini secin:")
    log(f"26SS: {ss_yolu.name}", "OK")

    hafta_str = datetime.now().strftime("%d.%m.%Y - Hafta %V")
    cikti_dir = SCRIPT_DIR / "raporlar" / datetime.now().strftime("%Y-W%V")
    cikti_dir.mkdir(parents=True, exist_ok=True)

    log("26AW Excel okunuyor...", ">>")
    df_aw = oku_excel(aw_yolu)
    log("26SS Excel okunuyor...", ">>")
    df_ss = oku_excel(ss_yolu)

    log("HTML olusturuluyor...", ">>")
    html = uret_html_cift(df_aw, df_ss, hafta_str, aw_yolu.name, ss_yolu.name)
    rapor = cikti_dir / f"lacoste_sezon_{datetime.now().strftime('%Y%m%d')}.html"
    with open(rapor, "w", encoding="utf-8-sig") as f:
        f.write(html)
    log(f"HTML: {rapor.name} ({len(html)//1024} KB)", "OK")

    if args.indir:
        for kanal in KANALLAR:
            excel_gitmeyen(df_aw, kanal, cikti_dir)

    if not args.no_ac:
        webbrowser.open(rapor.as_uri())
        log("Tarayicida acildi", "OK")

    print(f"\n{'='*55}")
    print(f"  Tamamlandi! {rapor}")
    print(f"{'='*55}\n")



if __name__ == "__main__":
    main()
